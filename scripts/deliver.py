#!/usr/bin/env python3
"""
deliver.py — retention-analysis Phase 5 (deliverable).

Reads compute.py JSON + the long-format CSV and writes a three-sheet
formula-driven Excel workbook.

Three sheet-layout modes:

* aggregating  — source has many rows per customer (one per product line / type).
                 Helper sheet = "Raw Data with Analysis" with SUMIFS aggregating by
                 customer + type filter. Includes the self-validation block at
                 the top (active counts, type-decomposed MRR totals, recon
                 against Raw Data direct sums).
* passthrough  — source already has one row per customer. Helper sheet =
                 "Raw Data with Analysis" with live 1:1 refs back to Raw Data.
* twotab       — no source workbook; Raw Data tab is built from the CSV.

Sheet 3 ("Raw Data") is ALWAYS a verbatim copy of the source workbook in
aggregating/passthrough modes — zero edits, zero reformatting, no color
changes. The skill's reconciliation guarantee depends on this.

Formatting follows the skill's finance convention:
  * Blue font (#0000FF)  — hardcoded inputs (ARR factor, methodology label values)
  * Green font (#006100) — formulas that reference another sheet
  * Black font           — formulas internal to the current sheet
  * $ symbol only on the top + bottom row of a vertical numeric block;
    interior rows use #,##0 with no $
  * Title fill #1F4E79, sub-header fill #D9E1F2, no merged cells (uses
    centerContinuous alignment)

CLI:
    python3 deliver.py <compute-output.json> <long-format-csv> <output.xlsx> \\
        [--company "<name>"] \\
        [--source <source.xlsx> --source-sheet "<sheet>" \\
         --source-customer-col <letter> --source-first-data-row <n> \\
         --source-first-date-col <letter>] \\
        [--source-type-col <letter> --type-filter "Recurring,Re-occurring"] \\
        [--lookback 12]

With --source-type-col, the aggregating mode is selected automatically.
With --source but no --source-type-col, passthrough mode is used.
With neither, two-tab fallback is used.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import sys
from copy import copy as _copy
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string


# ---------------------------------------------------------------------------
# Layout constants (Corkscrew)
# ---------------------------------------------------------------------------

ROW_TITLE = 1
ROW_GENERATED = 2
ROW_ARR_FACTOR = 3
ROW_DATES = 5
ROW_VS = 6  # optional "vs prior year" label row

# Rollforward block
ROW_BEGIN = 8
ROW_NEW = 9
ROW_UPSELL = 10
ROW_DOWNSELL = 11
ROW_CHURN = 12
ROW_END = 13
ROW_CHECK = 14

# Customer count block
ROW_CC_BANNER = 16
ROW_N_ACTIVE_PRIOR = 17
ROW_N_ACTIVE_CURR = 18
ROW_N_CHURNED = 19
ROW_N_NEW = 20

# Retention metrics
ROW_RR_BANNER = 22
ROW_GRR = 23
ROW_NRR = 24
ROW_LOGO = 25

# Per-customer metrics
ROW_PC_BANNER = 27
ROW_AVG_ARR = 28
ROW_AVG_NEW = 29

# Decomposed reconciliation (only when multi-type scope)
ROW_RECON_BANNER = 31
ROW_REC_RECURRING = 32
ROW_REC_REOCCURRING = 33
ROW_REC_SUM = 35
ROW_REC_VARIANCE = 36

COL_LABEL = 1
FIRST_DATA_COL = 2  # column B = first date column


# ---------------------------------------------------------------------------
# Layout constants (Raw Data with Analysis / helper)
# ---------------------------------------------------------------------------

ANALYSIS_ROW_HDR = 1
ANALYSIS_ROW_ACTIVE = 2
ANALYSIS_ROW_RETAINED = 3
ANALYSIS_ROW_CHECK_ACTIVE = 4
# row 5 blank divider
ANALYSIS_ROW_REC = 6
ANALYSIS_ROW_REOCC = 7
ANALYSIS_ROW_NONREC = 8
ANALYSIS_ROW_TOTAL = 9
ANALYSIS_ROW_CHECK_TOTAL = 10
ANALYSIS_ROW_CHECK_INSCOPE = 11
ANALYSIS_FIRST_CUST_ROW = 12

ANALYSIS_LABEL_COL = 1
ANALYSIS_FIRST_MONTH_COL = 2  # column B


# ---------------------------------------------------------------------------
# Colors & formats
# ---------------------------------------------------------------------------

TITLE_FILL = "1F4E79"      # dark blue, white bold text
BANNER_FILL = "1F4E79"     # section banners use same dark blue
SUBHEADER_FILL = "D9E1F2"  # light blue, black bold
KEY_METRIC_FILL = "BDD7EE" # medium blue for output rows (Ending, retention rates)

COLOR_BLUE = "0000FF"      # hardcoded inputs
COLOR_GREEN = "006100"     # cross-sheet references
COLOR_BLACK = "000000"     # formulas internal to current sheet
COLOR_WHITE = "FFFFFF"

# Number formats. Top-and-bottom-of-block rows get $; interior rows do not.
FMT_DOLLAR = '"$"#,##0;("$"#,##0);"-"'
FMT_NUMBER = '#,##0;(#,##0);"-"'
FMT_PCT = '0.0%;(0.0%);"-"'
FMT_COUNT = '#,##0;(#,##0);"-"'
FMT_DATE = "mmm-yy"


def font_hardcode(bold: bool = False, size: int = 10) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=COLOR_BLUE)


def font_xsheet(bold: bool = False, size: int = 10) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=COLOR_GREEN)


def font_formula(bold: bool = False, size: int = 10) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=COLOR_BLACK)


def font_title() -> Font:
    return Font(name="Calibri", size=14, bold=True, color=COLOR_WHITE)


def font_banner() -> Font:
    return Font(name="Calibri", size=10, bold=True, color=COLOR_WHITE)


def font_subheader() -> Font:
    return Font(name="Calibri", size=10, bold=True, color=COLOR_BLACK)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def center_continuous_across(ws, row: int, first_col: int, last_col: int,
                             text: str, font_obj: Font, fill_obj: PatternFill) -> None:
    """Write `text` to the leftmost cell and apply centerContinuous alignment
    across all cells in the span. Avoids merge_cells, which breaks
    selection/sort/filter/copy-paste."""
    for c in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.value = text if c == first_col else None
        cell.font = font_obj
        cell.fill = fill_obj
        cell.alignment = Alignment(horizontal="centerContinuous", vertical="center")


# ---------------------------------------------------------------------------
# Input loading
# ---------------------------------------------------------------------------


def load_compute_json(path: str) -> dict:
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


def load_long_csv(path: str) -> Tuple[List[str], List[str], Dict[Tuple[str, str], float]]:
    customers: set = set()
    months: set = set()
    cell: Dict[Tuple[str, str], float] = {}
    with open(path, "r", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            cust = str(row["customer_id"]).strip()
            month_raw = str(row["month"]).strip()
            month = month_raw[:7] if len(month_raw) >= 7 else month_raw
            try:
                mrr = float(row["mrr"])
            except (TypeError, ValueError):
                mrr = 0.0
            customers.add(cust)
            months.add(month)
            cell[(cust, month)] = mrr

    def cust_key(c: str):
        # Parse trailing integer if present (e.g. "Customer 178" -> 178)
        import re
        m = re.search(r"(\d+)\s*$", c)
        if m:
            return (0, int(m.group(1)), c)
        try:
            return (0, int(c), c)
        except ValueError:
            return (1, 0, c)

    customer_list = sorted(customers, key=cust_key)
    month_list = sorted(months)
    return customer_list, month_list, cell


def parse_month_cutoff(s: str) -> str:
    """Parse an --actuals-through value into a canonical 'YYYY-MM' string.
    Accepts 'YYYY-MM', 'YYYY-MM-DD', 'May-26', 'May 2026', '2026-M5', etc.
    (survey.py reports the cutoff as e.g. 'May-26')."""
    s = str(s).strip()
    for fmt in ("%Y-%m", "%Y-%m-%d", "%b-%y", "%b-%Y", "%B %Y", "%b %Y"):
        try:
            d = dt.datetime.strptime(s, fmt).date()
            return f"{d.year:04d}-{d.month:02d}"
        except ValueError:
            continue
    if "-M" in s:  # '2026-M5'
        try:
            y, mm = s.split("-M")
            return f"{int(y):04d}-{int(mm):02d}"
        except ValueError:
            pass
    raise ValueError(
        f"Could not parse --actuals-through {s!r}. Use 'YYYY-MM' (e.g. 2026-05) "
        f"or a month label like 'May-26'."
    )


def month_to_date(month_str: str) -> dt.date:
    y, m = month_str.split("-")[:2]
    return dt.date(int(y), int(m), 1)


def fmt_month_label(month_str: str) -> str:
    """e.g. '2021-01' -> '2021-M1' (matches example output style)."""
    y, m = month_str.split("-")[:2]
    return f"{int(y):04d}-M{int(m)}"


# ---------------------------------------------------------------------------
# Raw Data tab — verbatim copy of source
# ---------------------------------------------------------------------------


def copy_source_sheet_verbatim(src_path: str, src_ws, dest_ws) -> None:
    """Copy a source worksheet (or CSV) into dest_ws preserving values, number
    formats, fonts, fills, alignment, borders, merged ranges, column widths,
    row heights, and cell comments. Critical Rule 7: zero edits, no
    reformatting, no color changes."""
    if src_path.lower().endswith(".csv"):
        with open(src_path, "r", encoding="utf-8") as fh:
            reader = csv.reader(fh)
            for r_idx, row in enumerate(reader, start=1):
                for c_idx, val in enumerate(row, start=1):
                    cast: Any = val
                    if isinstance(val, str):
                        s = val.strip()
                        if s == "":
                            cast = None
                        else:
                            try:
                                cast = float(s) if "." in s else int(s)
                            except (TypeError, ValueError):
                                cast = val
                    dest_ws.cell(row=r_idx, column=c_idx, value=cast)
        return

    for row in src_ws.iter_rows():
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            dest_cell = dest_ws.cell(row=cell.row, column=cell.column,
                                     value=cell.value)
            if cell.has_style:
                dest_cell.font = _copy(cell.font)
                dest_cell.fill = _copy(cell.fill)
                dest_cell.border = _copy(cell.border)
                dest_cell.alignment = _copy(cell.alignment)
                dest_cell.number_format = cell.number_format
                dest_cell.protection = _copy(cell.protection)
            if cell.comment is not None:
                dest_cell.comment = Comment(cell.comment.text or "",
                                            cell.comment.author or "source")

    for merged_range in src_ws.merged_cells.ranges:
        dest_ws.merge_cells(str(merged_range))

    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width is not None:
            dest_ws.column_dimensions[col_letter].width = dim.width
    for row_num, dim in src_ws.row_dimensions.items():
        if dim.height is not None:
            dest_ws.row_dimensions[row_num].height = dim.height


# ---------------------------------------------------------------------------
# Two-tab Raw Data tab (no source — built from long CSV)
# ---------------------------------------------------------------------------


def write_raw_from_csv(ws, customers: List[str], months: List[str],
                       cell: Dict[Tuple[str, str], float]) -> None:
    """Two-tab fallback: build Raw Data sheet directly from the long CSV.
    Used only when no --source is supplied."""
    ws.cell(row=1, column=1, value="Customer ID").font = font_subheader()
    ws.cell(row=1, column=1).fill = fill(SUBHEADER_FILL)
    for j, m in enumerate(months):
        c = ws.cell(row=1, column=2 + j, value=month_to_date(m))
        c.number_format = FMT_DATE
        c.font = font_subheader()
        c.fill = fill(SUBHEADER_FILL)
        c.alignment = Alignment(horizontal="center")

    for i, cust in enumerate(customers):
        r = 2 + i
        ws.cell(row=r, column=1, value=cust).font = font_subheader()
        for j, m in enumerate(months):
            v = cell.get((cust, m), 0.0)
            c = ws.cell(row=r, column=2 + j, value=float(v))
            c.number_format = FMT_NUMBER
            c.font = font_hardcode()

    ws.column_dimensions["A"].width = 14
    for j in range(len(months)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 12
    ws.freeze_panes = ws.cell(row=2, column=2)


# ---------------------------------------------------------------------------
# Helper sheet — Raw Data with Analysis (aggregating mode)
# ---------------------------------------------------------------------------


def write_analysis_sheet(
    ws,
    customers: List[str],
    months_analysis: List[str],
    src_sheet_name: str,
    src_customer_col: str,
    src_type_col: str,
    src_first_data_row: int,
    src_last_data_row: int,
    src_first_date_col: str,
    in_scope_types: List[str],
    raw_sheet_name: str = "Raw Data",
) -> None:
    """Build the Raw Data with Analysis helper sheet.

    Layout (per the new SKILL.md):
      Row 1   Month headers          col A = "Customer ID"
      Row 2   # Active customers     COUNTIF on customer rows
      Row 3   # Retained vs N prior  SUMPRODUCT of two-month >0 masks (array)
      Row 4   Check # Active vs Raw  independent recount against Raw Data
      Row 5   blank divider
      Row 6   Recurring MRR total    SUMIFS direct column ref by type
      Row 7   Re-occurring MRR total
      Row 8   Non-recurring MRR total
      Row 9   Total MRR (all types)  = row6+row7+row8
      Row 10  Check vs Raw Data      direct SUM of column on Raw Data
      Row 11  Check (in-scope sum)   = in-scope rows − SUM(customer rows)
      Row 12+ Customer-level data    direct-column SUMIFS by customer + type

    All cells that pull from Raw Data are green (cross-sheet ref). Section
    label cells are black bold.
    """
    # CRITICAL: formulas reference the DESTINATION sheet name ("Raw Data"),
    # not the original source sheet name. The verbatim Raw Data tab keeps the
    # source's contents but is named "Raw Data" in this workbook.
    src_sheet_name = raw_sheet_name
    n_months = len(months_analysis)
    n_cust = len(customers)
    last_cust_row = ANALYSIS_FIRST_CUST_ROW + n_cust - 1

    src_first_date_idx = column_index_from_string(src_first_date_col)

    # Source helper strings — used in every formula
    type_rng = f"'{src_sheet_name}'!${src_type_col}${src_first_data_row}:${src_type_col}${src_last_data_row}"
    cust_rng = f"'{src_sheet_name}'!${src_customer_col}${src_first_data_row}:${src_customer_col}${src_last_data_row}"

    # Header row 1
    hdr_a = ws.cell(row=ANALYSIS_ROW_HDR, column=ANALYSIS_LABEL_COL, value="Customer ID")
    hdr_a.font = font_subheader()
    hdr_a.fill = fill(SUBHEADER_FILL)
    hdr_a.alignment = Alignment(horizontal="left")

    for j, m in enumerate(months_analysis):
        col = ANALYSIS_FIRST_MONTH_COL + j
        cell = ws.cell(row=ANALYSIS_ROW_HDR, column=col, value=fmt_month_label(m))
        cell.font = font_subheader()
        cell.fill = fill(SUBHEADER_FILL)
        cell.alignment = Alignment(horizontal="center")

    # Row 2: # Active customers (COUNTIF on customer rows of THIS sheet)
    ws.cell(row=ANALYSIS_ROW_ACTIVE, column=ANALYSIS_LABEL_COL,
            value="# Active customers").font = font_subheader()
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        col_letter = get_column_letter(col)
        f = f"=COUNTIF({col_letter}${ANALYSIS_FIRST_CUST_ROW}:{col_letter}${last_cust_row},\">0\")"
        c = ws.cell(row=ANALYSIS_ROW_ACTIVE, column=col, value=f)
        c.font = font_formula()
        c.number_format = FMT_COUNT

    # Row 3: # Retained vs N prior. For first <lookback> cols, value is "n/a".
    LOOKBACK = 12  # YoY; if the dataset is shorter the model is degenerate but harmless
    ws.cell(row=ANALYSIS_ROW_RETAINED, column=ANALYSIS_LABEL_COL,
            value=f"# Retained vs {LOOKBACK}mo prior").font = font_subheader()
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        col_letter = get_column_letter(col)
        if j < LOOKBACK:
            ws.cell(row=ANALYSIS_ROW_RETAINED, column=col, value="n/a").font = font_formula()
        else:
            prior_letter = get_column_letter(col - LOOKBACK)
            f = (f"=SUMPRODUCT(({col_letter}${ANALYSIS_FIRST_CUST_ROW}:{col_letter}${last_cust_row}>0)"
                 f"*({prior_letter}${ANALYSIS_FIRST_CUST_ROW}:{prior_letter}${last_cust_row}>0))")
            c = ws.cell(row=ANALYSIS_ROW_RETAINED, column=col, value=f)
            c.font = font_formula()
            c.number_format = FMT_COUNT

    # Row 4: Check # Active vs Raw Data — independent recount.
    # Uses SUMPRODUCT over the in-scope type filter against the customer list.
    ws.cell(row=ANALYSIS_ROW_CHECK_ACTIVE, column=ANALYSIS_LABEL_COL,
            value="  Check # Active vs Raw Data").font = font_formula()
    type_filter_or = ",".join(in_scope_types)
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        col_letter = get_column_letter(col)
        src_col_letter = get_column_letter(src_first_date_idx + j)
        src_rng = f"'{src_sheet_name}'!${src_col_letter}${src_first_data_row}:${src_col_letter}${src_last_data_row}"
        # An "active" customer is one with a positive in-scope sum.
        # The check is: COUNTIF on analysis sheet row above = count of customers with
        # positive sum on raw direct path.
        # For multi-type filter, use SUMPRODUCT(--(SUMIFS-array > 0)).
        # Since openpyxl can't easily write CSE array formulas, we use a
        # simpler equivalent: count customers where the in-scope SUM is > 0.
        # For single in-scope type — use COUNTIFS directly.
        if len(in_scope_types) == 1:
            t = in_scope_types[0]
            # COUNTIFS counts source rows matching type AND positive — but a
            # customer can have multiple positive rows (multi-product). To
            # count CUSTOMERS we need an aggregation. We approximate with
            # SUMPRODUCT against a unique-customer list. Simpler approach:
            # SUMPRODUCT(1/COUNTIFS) — but that fails on zero rows.
            # Use the array form via SUMPRODUCT with SUMIFS, which works in
            # Excel/LibreOffice as an implicit array context.
            f = (f"={col_letter}{ANALYSIS_ROW_ACTIVE}"
                 f" - SUMPRODUCT(--("
                 f"SUMIFS({src_rng},{type_rng},\"{t}\",{cust_rng},$A${ANALYSIS_FIRST_CUST_ROW}:$A${last_cust_row})>0))")
        else:
            # Multi-type: chain SUMIFS sums per type, then OR via sum.
            # SUMPRODUCT(--((SUMIFS_type1 + SUMIFS_type2 + ...) > 0))
            sumifs_parts = []
            for t in in_scope_types:
                sumifs_parts.append(
                    f"SUMIFS({src_rng},{type_rng},\"{t}\",{cust_rng},$A${ANALYSIS_FIRST_CUST_ROW}:$A${last_cust_row})"
                )
            inner = " + ".join(sumifs_parts)
            f = (f"={col_letter}{ANALYSIS_ROW_ACTIVE}"
                 f" - SUMPRODUCT(--(({inner}) > 0))")
        c = ws.cell(row=ANALYSIS_ROW_CHECK_ACTIVE, column=col, value=f)
        c.font = font_xsheet()  # cross-sheet, green
        c.number_format = FMT_COUNT

    # Row 5 — blank divider (intentionally empty)

    # Rows 6/7/8: per-type totals. Direct column reference per month — each
    # SUMIFS sums the source's date column for rows matching this type.
    type_rows = [
        (ANALYSIS_ROW_REC, "Recurring"),
        (ANALYSIS_ROW_REOCC, "Re-occurring"),
        (ANALYSIS_ROW_NONREC, "Non-recurring"),
    ]

    for row, type_name in type_rows:
        # Label in column A (black formula font)
        ws.cell(row=row, column=ANALYSIS_LABEL_COL,
                value=f"  {type_name}").font = font_formula()
        # If this type isn't in the source at all, still write the row but with
        # SUMIFS that yield 0; useful for the row 10 full-type recon.
        for j in range(n_months):
            col = ANALYSIS_FIRST_MONTH_COL + j
            src_col_letter = get_column_letter(src_first_date_idx + j)
            src_rng_j = (f"'{src_sheet_name}'!"
                         f"${src_col_letter}${src_first_data_row}:"
                         f"${src_col_letter}${src_last_data_row}")
            f = f"=SUMIFS({src_rng_j},{type_rng},\"{type_name}\")"
            c = ws.cell(row=row, column=col, value=f)
            c.font = font_xsheet()
            c.number_format = FMT_NUMBER

    # Row 9: Total MRR (all types)
    ws.cell(row=ANALYSIS_ROW_TOTAL, column=ANALYSIS_LABEL_COL,
            value="Total MRR (all types)").font = font_subheader()
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        col_letter = get_column_letter(col)
        f = (f"={col_letter}{ANALYSIS_ROW_REC}+{col_letter}{ANALYSIS_ROW_REOCC}"
             f"+{col_letter}{ANALYSIS_ROW_NONREC}")
        c = ws.cell(row=ANALYSIS_ROW_TOTAL, column=col, value=f)
        c.font = font_formula(bold=True)
        c.number_format = FMT_NUMBER

    # Row 10: Check vs Raw Data direct column sum (must = 0)
    ws.cell(row=ANALYSIS_ROW_CHECK_TOTAL, column=ANALYSIS_LABEL_COL,
            value="  Check vs Raw Data").font = font_formula()
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        col_letter = get_column_letter(col)
        src_col_letter = get_column_letter(src_first_date_idx + j)
        src_rng_j = (f"'{src_sheet_name}'!"
                     f"${src_col_letter}${src_first_data_row}:"
                     f"${src_col_letter}${src_last_data_row}")
        f = f"={col_letter}{ANALYSIS_ROW_TOTAL} - SUM({src_rng_j})"
        c = ws.cell(row=ANALYSIS_ROW_CHECK_TOTAL, column=col, value=f)
        c.font = font_xsheet()
        c.number_format = FMT_NUMBER

    # Row 11: Check (in-scope sum) — sum of in-scope type rows above − sum of
    # customer rows below; must = 0 if the customer-row SUMIFS uses the same
    # type filter.
    label = "  Check (" + " + ".join(in_scope_types) + ") vs customer rows"
    ws.cell(row=ANALYSIS_ROW_CHECK_INSCOPE, column=ANALYSIS_LABEL_COL,
            value=label).font = font_formula()
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        col_letter = get_column_letter(col)
        in_scope_sum = " + ".join(
            f"{col_letter}{ANALYSIS_ROW_REC if t == 'Recurring' else (ANALYSIS_ROW_REOCC if t == 'Re-occurring' else ANALYSIS_ROW_NONREC)}"
            for t in in_scope_types
        )
        f = (f"=({in_scope_sum}) - SUM({col_letter}${ANALYSIS_FIRST_CUST_ROW}:"
             f"{col_letter}${last_cust_row})")
        c = ws.cell(row=ANALYSIS_ROW_CHECK_INSCOPE, column=col, value=f)
        c.font = font_formula()
        c.number_format = FMT_NUMBER

    # Rows 12+: per-customer monthly in-scope MRR.
    # SUMIFS with two criteria: customer id + type filter (one row per in-scope
    # type, summed). When there's >1 in-scope type we add the SUMIFS terms.
    for i, cust in enumerate(customers):
        r = ANALYSIS_FIRST_CUST_ROW + i
        # Customer ID literal (or formula to source? — leave literal since the
        # source has many rows per customer)
        ws.cell(row=r, column=ANALYSIS_LABEL_COL, value=cust).font = font_subheader()
        for j in range(n_months):
            col = ANALYSIS_FIRST_MONTH_COL + j
            src_col_letter = get_column_letter(src_first_date_idx + j)
            src_rng_j = (f"'{src_sheet_name}'!"
                         f"${src_col_letter}${src_first_data_row}:"
                         f"${src_col_letter}${src_last_data_row}")
            sumifs_parts = []
            for t in in_scope_types:
                sumifs_parts.append(
                    f"SUMIFS({src_rng_j},{cust_rng},$A{r},{type_rng},\"{t}\")"
                )
            f = "=" + " + ".join(sumifs_parts)
            c = ws.cell(row=r, column=col, value=f)
            c.font = font_xsheet()
            c.number_format = FMT_NUMBER

    # Column widths
    ws.column_dimensions[get_column_letter(ANALYSIS_LABEL_COL)].width = 38
    for j in range(n_months):
        ws.column_dimensions[get_column_letter(ANALYSIS_FIRST_MONTH_COL + j)].width = 12

    ws.freeze_panes = ws.cell(row=ANALYSIS_FIRST_CUST_ROW, column=ANALYSIS_FIRST_MONTH_COL)


# ---------------------------------------------------------------------------
# Corkscrew sheet — aggregating mode (references Raw Data with Analysis)
# ---------------------------------------------------------------------------


def write_corkscrew_sheet_aggregating(
    ws,
    customers: List[str],
    months_analysis: List[str],
    arr_factor: float,
    company: str,
    in_scope_types: List[str],
    analysis_sheet_name: str,
    lookback: int = 12,
) -> None:
    """Write the Corkscrew with YoY rollforward (or N-period lookback).
    All movement formulas reference the Raw Data with Analysis prior and current
    columns. Multi-type recon block written when len(in_scope_types) > 1."""
    n_months = len(months_analysis)
    n_cust = len(customers)
    last_cust_row = ANALYSIS_FIRST_CUST_ROW + n_cust - 1
    n_periods = n_months - lookback  # number of comparison periods
    if n_periods <= 0:
        raise ValueError(
            f"Not enough months for {lookback}-period lookback: {n_months} months."
        )

    months_periods = months_analysis[lookback:]  # comparison-period labels

    # Title (row 1) — centerContinuous, navy fill, white bold
    title_text = (
        f"{company} — YoY ARR Corkscrew & Retention Analysis"
        if company else "YoY ARR Corkscrew & Retention Analysis"
    )
    center_continuous_across(
        ws, ROW_TITLE, 1, FIRST_DATA_COL + n_periods - 1,
        title_text, font_title(), fill(TITLE_FILL),
    )
    ws.row_dimensions[ROW_TITLE].height = 22

    # Generated row
    c = ws.cell(row=ROW_GENERATED, column=1, value="Generated:")
    c.font = font_subheader()
    c2 = ws.cell(row=ROW_GENERATED, column=2, value=dt.date.today().isoformat())
    c2.font = font_hardcode()

    # ARR factor (hardcode — blue)
    label = ws.cell(row=ROW_ARR_FACTOR, column=1, value="ARR Factor (MRR × N):")
    label.font = font_subheader()
    af = ws.cell(row=ROW_ARR_FACTOR, column=2, value=int(arr_factor))
    af.font = font_hardcode(bold=True)
    af.number_format = "0"
    af.comment = Comment(
        f"Source: User-confirmed in Phase 1. Data is "
        f"{'MRR (annualize ×12)' if int(arr_factor) == 12 else 'ARR (factor 1)'}.",
        "retention-analysis"
    )
    arr_ref = f"$B${ROW_ARR_FACTOR}"

    # Date row 5
    lbl5 = ws.cell(row=ROW_DATES, column=COL_LABEL, value="Item")
    lbl5.font = font_banner()
    lbl5.fill = fill(BANNER_FILL)
    lbl5.alignment = Alignment(horizontal="left")
    for j, m in enumerate(months_periods):
        col = FIRST_DATA_COL + j
        c = ws.cell(row=ROW_DATES, column=col, value=fmt_month_label(m))
        c.font = font_banner()
        c.fill = fill(BANNER_FILL)
        c.alignment = Alignment(horizontal="center")

    # Row 6: "(vs. prior year)" prefix
    vs_label = ws.cell(row=ROW_VS, column=COL_LABEL, value="(vs. prior year)")
    vs_label.font = font_subheader()
    vs_label.fill = fill(SUBHEADER_FILL)
    vs_label.alignment = Alignment(horizontal="left")
    for j, m in enumerate(months_periods):
        col = FIRST_DATA_COL + j
        prior_label = fmt_month_label(months_analysis[j])  # prior = T-N
        c = ws.cell(row=ROW_VS, column=col, value=f"vs {prior_label}")
        c.font = font_subheader()
        c.fill = fill(SUBHEADER_FILL)
        c.alignment = Alignment(horizontal="center")

    # Rollforward block. Exactly ONE external check is emitted per column:
    #   - single in-scope type  → ROW_CHECK (row 14) = Ending − period total summed
    #     independently from the analysis/Raw Data sheet × factor.
    #   - multiple in-scope types → the variance at the bottom of the decomposed
    #     reconciliation block (row 36) is the check, with components shown above it.
    # DRY: one check, never two paths to the same algebra.
    rollforward_labels = {
        ROW_BEGIN: ("Beginning ARR (prior year)", True),     # top-of-block: $
        ROW_NEW: ("  + New customer ARR", False),
        ROW_UPSELL: ("  + Expansion (Upsell)", False),
        ROW_DOWNSELL: ("  - Contraction (Downsell)", False),
        ROW_CHURN: ("  - Churn", False),
        ROW_END: ("Ending ARR", True),                        # bottom-of-block: $
    }
    for r, (lbl, _) in rollforward_labels.items():
        c = ws.cell(row=r, column=COL_LABEL, value=lbl)
        c.font = font_subheader()
        c.alignment = Alignment(horizontal="left")
        if r in (ROW_BEGIN, ROW_END):
            c.fill = fill(KEY_METRIC_FILL)

    # Helper: analysis sheet column letter for a given source month index
    def analysis_col(month_index: int) -> str:
        return get_column_letter(ANALYSIS_FIRST_MONTH_COL + month_index)

    for j in range(n_periods):
        col = FIRST_DATA_COL + j
        col_letter = get_column_letter(col)
        curr_idx = lookback + j           # analysis sheet column for current period
        prior_idx = j                     # analysis sheet column for prior period
        curr = analysis_col(curr_idx)
        prior = analysis_col(prior_idx)
        # Analysis sheet data ranges
        rc = f"'{analysis_sheet_name}'!{curr}${ANALYSIS_FIRST_CUST_ROW}:{curr}${last_cust_row}"
        rp = f"'{analysis_sheet_name}'!{prior}${ANALYSIS_FIRST_CUST_ROW}:{prior}${last_cust_row}"

        # Beginning ARR = SUMPRODUCT((prior > 0) * prior) * ARR_factor
        f_beg = f"=SUMPRODUCT(({rp}>0)*{rp})*{arr_ref}"
        c = ws.cell(row=ROW_BEGIN, column=col, value=f_beg)
        c.font = font_xsheet(bold=True)
        c.number_format = FMT_DOLLAR  # top of block — $
        c.fill = fill(KEY_METRIC_FILL)

        # New = SUMPRODUCT((prior=0)*(curr>0)*curr) * factor
        f_new = f"=SUMPRODUCT(({rp}=0)*({rc}>0)*{rc})*{arr_ref}"
        c = ws.cell(row=ROW_NEW, column=col, value=f_new)
        c.font = font_xsheet()
        c.number_format = FMT_NUMBER  # interior — no $

        # Upsell = SUMPRODUCT((prior>0)*(curr>prior)*(curr-prior)) * factor
        f_up = f"=SUMPRODUCT(({rp}>0)*({rc}>{rp})*({rc}-{rp}))*{arr_ref}"
        c = ws.cell(row=ROW_UPSELL, column=col, value=f_up)
        c.font = font_xsheet()
        c.number_format = FMT_NUMBER

        # Downsell (negative) = SUMPRODUCT((prior>0)*(curr>0)*(curr<prior)*(curr-prior)) * factor
        f_down = f"=SUMPRODUCT(({rp}>0)*({rc}>0)*({rc}<{rp})*({rc}-{rp}))*{arr_ref}"
        c = ws.cell(row=ROW_DOWNSELL, column=col, value=f_down)
        c.font = font_xsheet()
        c.number_format = FMT_NUMBER

        # Churn (negative) = SUMPRODUCT((prior>0)*(curr=0)*(-prior)) * factor
        f_ch = f"=SUMPRODUCT(({rp}>0)*({rc}=0)*(-{rp}))*{arr_ref}"
        c = ws.cell(row=ROW_CHURN, column=col, value=f_ch)
        c.font = font_xsheet()
        c.number_format = FMT_NUMBER

        # Ending = sum of rollforward
        f_end = (f"={col_letter}{ROW_BEGIN}+{col_letter}{ROW_NEW}"
                 f"+{col_letter}{ROW_UPSELL}+{col_letter}{ROW_DOWNSELL}"
                 f"+{col_letter}{ROW_CHURN}")
        c = ws.cell(row=ROW_END, column=col, value=f_end)
        c.font = font_formula(bold=True)
        c.number_format = FMT_DOLLAR  # bottom of block — $
        c.fill = fill(KEY_METRIC_FILL)

        # Customer counts via HLOOKUP into analysis sheet row 2
        analysis_hdr_range = (f"'{analysis_sheet_name}'!"
                          f"$B${ANALYSIS_ROW_HDR}:${get_column_letter(ANALYSIS_FIRST_MONTH_COL + n_months - 1)}${ANALYSIS_ROW_ACTIVE}")
        f_n_prior = f"=HLOOKUP(SUBSTITUTE({col_letter}${ROW_VS},\"vs \",\"\"),{analysis_hdr_range},2,FALSE)"
        f_n_curr = f"=HLOOKUP({col_letter}${ROW_DATES},{analysis_hdr_range},2,FALSE)"
        ws.cell(row=ROW_N_ACTIVE_PRIOR, column=col, value=f_n_prior).font = font_xsheet()
        ws.cell(row=ROW_N_ACTIVE_PRIOR, column=col).number_format = FMT_COUNT
        ws.cell(row=ROW_N_ACTIVE_CURR, column=col, value=f_n_curr).font = font_xsheet()
        ws.cell(row=ROW_N_ACTIVE_CURR, column=col).number_format = FMT_COUNT

        # Retained — pull from analysis sheet row 3
        analysis_ret_range = (f"'{analysis_sheet_name}'!"
                          f"$B${ANALYSIS_ROW_HDR}:${get_column_letter(ANALYSIS_FIRST_MONTH_COL + n_months - 1)}${ANALYSIS_ROW_RETAINED}")
        f_retained = f"=HLOOKUP({col_letter}${ROW_DATES},{analysis_ret_range},3,FALSE)"
        f_n_ch = f"={col_letter}{ROW_N_ACTIVE_PRIOR} - {f_retained[1:]}"  # = prior − retained
        ws.cell(row=ROW_N_CHURNED, column=col, value=f_n_ch).font = font_formula()
        ws.cell(row=ROW_N_CHURNED, column=col).number_format = FMT_COUNT
        f_n_new = f"={col_letter}{ROW_N_ACTIVE_CURR} - {f_retained[1:]}"  # = current − retained
        ws.cell(row=ROW_N_NEW, column=col, value=f_n_new).font = font_formula()
        ws.cell(row=ROW_N_NEW, column=col).number_format = FMT_COUNT

        # Retention metrics (with IFERROR)
        beg = f"{col_letter}{ROW_BEGIN}"
        f_grr = f"=IFERROR(({beg}+{col_letter}{ROW_DOWNSELL}+{col_letter}{ROW_CHURN})/{beg},0)"
        f_nrr = f"=IFERROR(({beg}+{col_letter}{ROW_UPSELL}+{col_letter}{ROW_DOWNSELL}+{col_letter}{ROW_CHURN})/{beg},0)"
        f_logo = f"=IFERROR(({col_letter}{ROW_N_ACTIVE_PRIOR}-{col_letter}{ROW_N_CHURNED})/{col_letter}{ROW_N_ACTIVE_PRIOR},0)"
        for r, fx in ((ROW_GRR, f_grr), (ROW_NRR, f_nrr), (ROW_LOGO, f_logo)):
            cc = ws.cell(row=r, column=col, value=fx)
            cc.font = font_formula()
            cc.number_format = FMT_PCT

        # Per-customer metrics
        f_avg = f"=IFERROR({col_letter}{ROW_END}/{col_letter}{ROW_N_ACTIVE_CURR},0)"
        f_avg_new = f"=IFERROR({col_letter}{ROW_NEW}/{col_letter}{ROW_N_NEW},0)"
        cc = ws.cell(row=ROW_AVG_ARR, column=col, value=f_avg)
        cc.font = font_formula()
        cc.number_format = FMT_DOLLAR  # singleton numeric row — top & bottom $
        cc = ws.cell(row=ROW_AVG_NEW, column=col, value=f_avg_new)
        cc.font = font_formula()
        cc.number_format = FMT_DOLLAR

        # Decomposed reconciliation (only when multi-type scope)
        if len(in_scope_types) > 1:
            for t in in_scope_types:
                if t == "Recurring":
                    f_rec = f"='{analysis_sheet_name}'!{curr}{ANALYSIS_ROW_REC}*{arr_ref}"
                    c = ws.cell(row=ROW_REC_RECURRING, column=col, value=f_rec)
                    c.font = font_xsheet()
                    c.number_format = FMT_DOLLAR
                elif t == "Re-occurring":
                    f_reocc = f"='{analysis_sheet_name}'!{curr}{ANALYSIS_ROW_REOCC}*{arr_ref}"
                    c = ws.cell(row=ROW_REC_REOCCURRING, column=col, value=f_reocc)
                    c.font = font_xsheet()
                    c.number_format = FMT_NUMBER
            f_sum = f"={col_letter}{ROW_REC_RECURRING}+{col_letter}{ROW_REC_REOCCURRING}"
            c = ws.cell(row=ROW_REC_SUM, column=col, value=f_sum)
            c.font = font_formula(bold=True)
            c.number_format = FMT_DOLLAR
            f_var = f"={col_letter}{ROW_REC_SUM}-{col_letter}{ROW_END}"
            c = ws.cell(row=ROW_REC_VARIANCE, column=col, value=f_var)
            c.font = font_formula()
            c.number_format = FMT_NUMBER
        else:
            # Single in-scope type → external reconciliation right under Ending.
            # Ending must equal the period's in-scope total summed independently
            # from the analysis/Raw Data sheet (a different formula path → a real
            # check, not the tautological Beginning+moves=Ending identity).
            f_chk = f"={col_letter}{ROW_END} - SUM({rc})*{arr_ref}"
            c = ws.cell(row=ROW_CHECK, column=col, value=f_chk)
            c.font = font_xsheet()
            c.number_format = FMT_NUMBER

    # Section banners (banner row above each block)
    banners = [
        (ROW_CC_BANNER, "CUSTOMER COUNTS", n_periods),
        (ROW_RR_BANNER, "RETENTION RATES", n_periods),
        (ROW_PC_BANNER, "PER-CUSTOMER METRICS", n_periods),
    ]
    if len(in_scope_types) > 1:
        banners.append((ROW_RECON_BANNER, "RECONCILIATION CHECKS", n_periods))
    for row, txt, span_cols in banners:
        center_continuous_across(
            ws, row, 1, FIRST_DATA_COL + span_cols - 1,
            txt, font_banner(), fill(BANNER_FILL),
        )

    # Row labels for CC / RR / PC / Recon sections
    rr_labels = {
        ROW_N_ACTIVE_PRIOR: "# Active (prior period)",
        ROW_N_ACTIVE_CURR: "# Active (current period)",
        ROW_N_CHURNED: "# Churned",
        ROW_N_NEW: "# New",
        ROW_GRR: "Gross Dollar Retention (GRR)",
        ROW_NRR: "Net Dollar Retention (NRR)",
        ROW_LOGO: "Logo Retention",
        ROW_AVG_ARR: "Avg ARR per Active Customer",
        ROW_AVG_NEW: "Avg ARR per New Customer",
    }
    if len(in_scope_types) > 1:
        rr_labels.update({
            ROW_REC_RECURRING: "Recurring ARR",
            ROW_REC_REOCCURRING: "Re-occurring ARR",
            ROW_REC_SUM: "Sum customer ARR",
            ROW_REC_VARIANCE: "Variance vs Ending ARR (= 0)",
        })
    else:
        rr_labels[ROW_CHECK] = "External Check (Ending - Raw Data) = 0"
    for r, txt in rr_labels.items():
        c = ws.cell(row=r, column=COL_LABEL, value=txt)
        c.font = font_subheader()
        c.alignment = Alignment(horizontal="left")

    # Column widths
    ws.column_dimensions[get_column_letter(COL_LABEL)].width = 38
    for j in range(n_periods):
        ws.column_dimensions[get_column_letter(FIRST_DATA_COL + j)].width = 13

    ws.freeze_panes = ws.cell(row=ROW_DATES + 2, column=FIRST_DATA_COL)


# ---------------------------------------------------------------------------
# Pass-through "Raw Data with Analysis" sheet — legacy (one row per customer source)
# ---------------------------------------------------------------------------


def write_analysis_passthrough_sheet(
    ws,
    customers: List[str],
    months: List[str],
    customer_to_src_row: Dict[str, int],
    src_customer_col: str,
    src_first_date_col: str,
    excluded_customers: List[str] | None = None,
) -> None:
    """Pass-through helper: one row per customer, formulas pulling Raw Data
    cells one-to-one. Used when the source already has one row per customer."""
    excluded_customers = excluded_customers or []
    src_date_col_idx = column_index_from_string(src_first_date_col)
    n_months = len(months)
    excl_col = ANALYSIS_FIRST_MONTH_COL + n_months
    last_cust_row = ANALYSIS_FIRST_CUST_ROW + len(customers) - 1

    # ---- Summary block (rows 1-3) — same shape the Corkscrew HLOOKUPs expect.
    # Row 1: string month headers ("2022-M1") so HLOOKUP matches the Corkscrew
    #        date row (also string labels). Row 2: # Active. Row 3: # Retained.
    hdr_a = ws.cell(row=ANALYSIS_ROW_HDR, column=ANALYSIS_LABEL_COL, value="Customer ID")
    hdr_a.font = font_subheader(); hdr_a.fill = fill(SUBHEADER_FILL)
    hdr_a.alignment = Alignment(horizontal="left")
    for j, m in enumerate(months):
        c = ws.cell(row=ANALYSIS_ROW_HDR, column=ANALYSIS_FIRST_MONTH_COL + j,
                    value=fmt_month_label(m))
        c.font = font_subheader(); c.fill = fill(SUBHEADER_FILL)
        c.alignment = Alignment(horizontal="center")
    ex_hdr = ws.cell(row=ANALYSIS_ROW_HDR, column=excl_col, value="Excluded?")
    ex_hdr.font = font_subheader(); ex_hdr.fill = fill(SUBHEADER_FILL)

    # Row 2: # Active customers (COUNTIF on this sheet's customer rows)
    ws.cell(row=ANALYSIS_ROW_ACTIVE, column=ANALYSIS_LABEL_COL,
            value="# Active customers").font = font_subheader()
    for j in range(n_months):
        cl = get_column_letter(ANALYSIS_FIRST_MONTH_COL + j)
        c = ws.cell(row=ANALYSIS_ROW_ACTIVE, column=ANALYSIS_FIRST_MONTH_COL + j,
                    value=f"=COUNTIF({cl}${ANALYSIS_FIRST_CUST_ROW}:{cl}${last_cust_row},\">0\")")
        c.font = font_formula(); c.number_format = FMT_COUNT

    # Row 3: # Retained vs N-mo prior ("n/a" for the first <lookback> columns)
    LOOKBACK = 12
    ws.cell(row=ANALYSIS_ROW_RETAINED, column=ANALYSIS_LABEL_COL,
            value=f"# Retained vs {LOOKBACK}mo prior").font = font_subheader()
    for j in range(n_months):
        col = ANALYSIS_FIRST_MONTH_COL + j
        cl = get_column_letter(col)
        if j < LOOKBACK:
            ws.cell(row=ANALYSIS_ROW_RETAINED, column=col, value="n/a").font = font_formula()
        else:
            pl = get_column_letter(col - LOOKBACK)
            c = ws.cell(row=ANALYSIS_ROW_RETAINED, column=col,
                        value=(f"=SUMPRODUCT(({cl}${ANALYSIS_FIRST_CUST_ROW}:{cl}${last_cust_row}>0)"
                               f"*({pl}${ANALYSIS_FIRST_CUST_ROW}:{pl}${last_cust_row}>0))"))
            c.font = font_formula(); c.number_format = FMT_COUNT

    first_data_row = ANALYSIS_FIRST_CUST_ROW

    for i, cust in enumerate(customers):
        r = first_data_row + i
        src_row = customer_to_src_row.get(str(cust))
        if src_row is not None:
            ws.cell(row=r, column=1, value=f"='Raw Data'!{src_customer_col}{src_row}").font = font_xsheet()
        else:
            ws.cell(row=r, column=1, value=cust).font = font_subheader()
        for j in range(n_months):
            col = ANALYSIS_FIRST_MONTH_COL + j
            if src_row is not None:
                src_col_letter = get_column_letter(src_date_col_idx + j)
                f = (f"=IF('Raw Data'!{src_col_letter}{src_row}=\"\",0,"
                     f"'Raw Data'!{src_col_letter}{src_row})")
                c = ws.cell(row=r, column=col, value=f)
                c.font = font_xsheet()
            else:
                c = ws.cell(row=r, column=col, value=0.0)
                c.font = font_hardcode()
            c.number_format = FMT_NUMBER
        ws.cell(row=r, column=excl_col, value=False).font = font_formula()

    # Excluded rows below
    base = first_data_row + len(customers)
    for i, cust in enumerate(excluded_customers):
        r = base + i
        src_row = customer_to_src_row.get(str(cust))
        if src_row is not None:
            ws.cell(row=r, column=1, value=f"='Raw Data'!{src_customer_col}{src_row}").font = font_xsheet()
        else:
            ws.cell(row=r, column=1, value=cust).font = font_subheader()
        for j in range(n_months):
            col = ANALYSIS_FIRST_MONTH_COL + j
            c = ws.cell(row=r, column=col, value=0.0)
            c.font = font_hardcode()
            c.number_format = FMT_NUMBER
        c2 = ws.cell(row=r, column=excl_col, value=True)
        c2.font = Font(name="Calibri", size=10, bold=True, color="C00000")

    ws.column_dimensions["A"].width = 18
    for j in range(n_months):
        ws.column_dimensions[get_column_letter(ANALYSIS_FIRST_MONTH_COL + j)].width = 12
    ws.column_dimensions[get_column_letter(excl_col)].width = 12
    ws.freeze_panes = ws.cell(row=first_data_row, column=ANALYSIS_FIRST_MONTH_COL)


def build_customer_to_src_row_map(src_path: str, src_customer_col: str,
                                  src_first_data_row: int, src_ws=None,
                                  src_last_data_row: int | None = None) -> Dict[str, int]:
    if src_path.lower().endswith(".csv"):
        out: Dict[str, int] = {}
        with open(src_path, "r", encoding="utf-8") as fh:
            reader = csv.DictReader(fh)
            seen = set()
            r = src_first_data_row
            for row in reader:
                cid = str(row.get("customer_id", "")).strip()
                if cid and cid not in seen:
                    out[cid] = r
                    seen.add(cid)
                    r += 1
        return out

    col_idx = column_index_from_string(src_customer_col)
    # Stop at src_last_data_row when given, so a summary/total block BELOW the
    # customer list (within the sheet) isn't scooped up as bogus "customers".
    last_row = src_last_data_row if src_last_data_row else src_ws.max_row
    out = {}
    for r in range(src_first_data_row, last_row + 1):
        v = src_ws.cell(row=r, column=col_idx).value
        if v is None or v == "":
            continue
        out[str(v).strip()] = r
    return out


# ---------------------------------------------------------------------------
# Source-sheet introspection helpers (aggregating mode)
# ---------------------------------------------------------------------------


def find_source_last_data_row(src_ws, src_customer_col: str,
                              src_first_data_row: int) -> int:
    """Find the last row in the source where the customer column has a value."""
    col_idx = column_index_from_string(src_customer_col)
    last = src_first_data_row
    for r in range(src_first_data_row, src_ws.max_row + 1):
        if src_ws.cell(row=r, column=col_idx).value not in (None, ""):
            last = r
    return last


def get_source_months(src_ws, src_first_date_col: str,
                      header_row: int = None) -> List[str]:
    """Read the date headers from the source sheet (the row above first data
    row, or an explicit header_row). Returns 'YYYY-MM' strings."""
    ws = src_ws
    first_idx = column_index_from_string(src_first_date_col)
    # If header_row not given, try row 1 then row 2.
    rows_to_try = [header_row] if header_row else [1, 2]
    for hr in rows_to_try:
        if hr is None or hr < 1:
            continue
        months = []
        c = first_idx
        while True:
            v = ws.cell(row=hr, column=c).value
            if v is None:
                break
            if isinstance(v, (dt.datetime, dt.date)):
                months.append(f"{v.year:04d}-{v.month:02d}")
            else:
                s = str(v).strip()
                # try parse
                parsed = None
                for fmt in ("%Y-%m-%d", "%Y-%m", "%b-%y", "%b %Y", "%B %Y", "%Y-M%m"):
                    try:
                        parsed = dt.datetime.strptime(s, fmt).date()
                        break
                    except ValueError:
                        continue
                if parsed is None and "-M" in s:
                    # "2021-M1" style
                    try:
                        y, mm = s.split("-M")
                        parsed = dt.date(int(y), int(mm), 1)
                    except ValueError:
                        pass
                if parsed is None:
                    break
                months.append(f"{parsed.year:04d}-{parsed.month:02d}")
            c += 1
        if len(months) >= 2:
            return months
    return []


# ---------------------------------------------------------------------------
# Top-level deliver()
# ---------------------------------------------------------------------------


def deliver(
    long_csv_path: str,
    output_xlsx_path: str,
    arr_factor: float = 12.0,
    compute_json_path: str | None = None,
    company: str = "",
    source_path: str | None = None,
    source_sheet: str | None = None,
    source_customer_col: str = "A",
    source_first_data_row: int = 2,
    source_first_date_col: str = "B",
    source_type_col: str | None = None,
    type_filter: List[str] | None = None,
    lookback: int = 12,
    source_header_row: int | None = None,
    source_last_data_row: int | None = None,
    actuals_through: str | None = None,
) -> str:
    customers, months, cell = load_long_csv(long_csv_path)

    # Actuals cutoff (#2): drop the in-progress current month and any forecast
    # tail so projections aren't counted as retention. Applies to every mode —
    # the helper and corkscrew iterate this month list.
    if actuals_through:
        cutoff = parse_month_cutoff(actuals_through)
        kept = [m for m in months if m <= cutoff]
        if len(kept) < 2:
            raise ValueError(
                f"--actuals-through {actuals_through!r} (= {cutoff}) leaves only "
                f"{len(kept)} month(s) of the {len(months)} in the CSV — need >= 2."
            )
        months = kept
    # compute.py output is OPTIONAL. The workbook is built independently from the
    # CSV as live formulas; if a compute.json is supplied we only adopt its
    # arr_factor (the agent can use the rest as an out-of-band cross-check).
    arr_factor = float(arr_factor)
    if compute_json_path:
        compute = load_compute_json(compute_json_path)
        arr_factor = float(compute.get("config", {}).get("arr_factor", arr_factor))

    wb = Workbook()
    ws_cork = wb.active
    ws_cork.title = "Corkscrew"

    mode = "twotab"
    if source_path and source_type_col:
        mode = "aggregating"
    elif source_path:
        mode = "passthrough"

    # Load the source workbook ONCE (was previously re-read up to 4x). Reused
    # for the verbatim Raw Data copy and every layout-discovery scan below.
    src_ws = None
    if source_path and not source_path.lower().endswith(".csv"):
        _src_wb = load_workbook(source_path, data_only=True, read_only=False)
        if source_sheet is None or source_sheet not in _src_wb.sheetnames:
            raise ValueError(
                f"Source sheet {source_sheet!r} not found in {source_path}. "
                f"Available: {_src_wb.sheetnames}"
            )
        src_ws = _src_wb[source_sheet]

    if mode == "aggregating":
        # Three sheets: Corkscrew, Raw Data with Analysis, Raw Data
        ws_analysis = wb.create_sheet("Raw Data with Analysis")
        ws_raw = wb.create_sheet("Raw Data")

        # 1. Raw Data — verbatim
        copy_source_sheet_verbatim(source_path, src_ws, ws_raw)

        # 2. Discover source layout. An explicit --source-last-data-row wins over
        # auto-detection (#3) — caps the customer block so a summary/total block
        # below it isn't aggregated in.
        src_last_row = source_last_data_row or find_source_last_data_row(
            src_ws, source_customer_col, source_first_data_row
        )
        in_scope = type_filter or ["Recurring", "Re-occurring"]
        # Use the analysis sheet's month list = full source month range
        months_analysis = get_source_months(
            src_ws, source_first_date_col, header_row=source_header_row,
        )
        if not months_analysis:
            # Fallback to the long-CSV month list
            months_analysis = months
        # Honor the actuals cutoff (#2) on the source-derived month list too.
        if actuals_through:
            months_analysis = [m for m in months_analysis if m <= parse_month_cutoff(actuals_through)]

        # 3. Raw Data with Analysis
        write_analysis_sheet(
            ws_analysis,
            customers=customers,
            months_analysis=months_analysis,
            src_sheet_name=source_sheet,
            src_customer_col=source_customer_col,
            src_type_col=source_type_col,
            src_first_data_row=source_first_data_row,
            src_last_data_row=src_last_row,
            src_first_date_col=source_first_date_col,
            in_scope_types=in_scope,
        )

        # 4. Corkscrew
        write_corkscrew_sheet_aggregating(
            ws_cork,
            customers=customers,
            months_analysis=months_analysis,
            arr_factor=arr_factor,
            company=company,
            in_scope_types=in_scope,
            analysis_sheet_name="Raw Data with Analysis",
            lookback=lookback,
        )

    elif mode == "passthrough":
        ws_helper = wb.create_sheet("Raw Data with Analysis")
        ws_raw = wb.create_sheet("Raw Data")
        copy_source_sheet_verbatim(source_path, src_ws, ws_raw)

        customer_to_src_row = build_customer_to_src_row_map(
            source_path, source_customer_col, source_first_data_row, src_ws,
            src_last_data_row=source_last_data_row,
        )
        src_customers_in_order = list(customer_to_src_row.keys())
        post_excl = set(map(str, customers))
        excluded = [c for c in src_customers_in_order if c not in post_excl]

        write_analysis_passthrough_sheet(
            ws_helper, customers, months,
            customer_to_src_row=customer_to_src_row,
            src_customer_col=source_customer_col,
            src_first_date_col=source_first_date_col,
            excluded_customers=excluded,
        )
        # Corkscrew references the pass-through helper — same formulas as
        # aggregating but pointing at a different helper sheet, with single-
        # type recon (passthrough = one type in scope by definition).
        write_corkscrew_sheet_aggregating(
            ws_cork,
            customers=customers,
            months_analysis=months,
            arr_factor=arr_factor,
            company=company,
            in_scope_types=["Recurring"],  # treated as a single-type bucket
            analysis_sheet_name="Raw Data with Analysis",
            lookback=lookback,
        )

    else:
        # twotab
        ws_raw = wb.create_sheet("Raw Data")
        write_raw_from_csv(ws_raw, customers, months, cell)
        write_corkscrew_sheet_aggregating(
            ws_cork,
            customers=customers,
            months_analysis=months,
            arr_factor=arr_factor,
            company=company,
            in_scope_types=["Recurring"],
            analysis_sheet_name="Raw Data",
            lookback=lookback,
        )

    wb.save(output_xlsx_path)
    return output_xlsx_path


def parse_args(argv):
    p = argparse.ArgumentParser(description="Retention-analysis Phase 5.")
    p.add_argument("long_csv")
    p.add_argument("output_xlsx")
    p.add_argument("--arr-factor", type=float, default=12.0,
                   help="MRR->ARR multiplier (12 for MRR input, 1 for ARR input)")
    p.add_argument("--compute-json", default=None,
                   help="OPTIONAL compute.py output. Not required — deliver builds "
                        "independently from the CSV. If given, its arr_factor "
                        "overrides --arr-factor.")
    p.add_argument("--company", default="")
    p.add_argument("--config", default=None,
                   help="JSON written by survey.py --emit-config. Fills the source-* "
                        "options and --actuals-through so survey's findings flow "
                        "straight in; any explicit flag still overrides it.")
    p.add_argument("--source", default=None)
    p.add_argument("--source-sheet", default=None)
    p.add_argument("--source-customer-col", default="A")
    p.add_argument("--source-first-data-row", type=int, default=2)
    p.add_argument("--source-last-data-row", type=int, default=None,
                   help="Last row of the customer block in the source (#3). Caps the "
                        "scan so a summary/total block below the customers isn't "
                        "treated as customers. Default: scan to the last non-empty row.")
    p.add_argument("--actuals-through", default=None,
                   help="Last COMPLETE actual month, e.g. '2026-05' or 'May-26' (#2). "
                        "Drops the in-progress current month and any forecast tail so "
                        "projections aren't counted as retention. Feed survey.py's "
                        "actuals_through here.")
    p.add_argument("--source-first-date-col", default="B")
    p.add_argument("--source-header-row", type=int, default=None,
                   help="Explicit row number of the date-header row "
                        "(default: try row above first-data-row)")
    p.add_argument("--source-type-col", default=None,
                   help="Column letter of revenue-type column (triggers "
                        "aggregating mode)")
    p.add_argument("--type-filter", default=None,
                   help="Comma-separated list of in-scope types "
                        "(default: 'Recurring,Re-occurring')")
    p.add_argument("--lookback", type=int, default=12)
    # Apply survey.py --emit-config values as defaults; explicit CLI flags win.
    pre, _ = p.parse_known_args(argv)
    if pre.config:
        with open(pre.config, "r", encoding="utf-8") as fh:
            cfg = json.load(fh)
        p.set_defaults(**{k: v for k, v in cfg.items() if v is not None})
    return p.parse_args(argv)


def main(argv=None):
    import time
    _t0 = time.perf_counter()
    args = parse_args(argv if argv is not None else sys.argv[1:])
    type_filter = (
        [t.strip() for t in args.type_filter.split(",")]
        if args.type_filter else None
    )
    out_xlsx = deliver(
        args.long_csv, args.output_xlsx,
        arr_factor=args.arr_factor,
        compute_json_path=args.compute_json,
        company=args.company,
        source_path=args.source,
        source_sheet=args.source_sheet,
        source_customer_col=args.source_customer_col,
        source_first_data_row=args.source_first_data_row,
        source_first_date_col=args.source_first_date_col,
        source_type_col=args.source_type_col,
        type_filter=type_filter,
        lookback=args.lookback,
        source_header_row=args.source_header_row,
        source_last_data_row=args.source_last_data_row,
        actuals_through=args.actuals_through,
    )
    print(f"Wrote: {out_xlsx}")
    print(f"[deliver.py] built workbook in {time.perf_counter() - _t0:.2f}s", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
