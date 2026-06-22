#!/usr/bin/env python3
"""
survey.py — retention-analysis Phases 1-4 helper.

Surveys an .xlsx or .csv workbook and produces a structured hypothesis report:
sheet roles, candidate customer column, candidate date columns, MRR-vs-ARR scale
signal, summary/total rows, derived blocks, negative values, and overall
sufficiency assessment.

This script does NOT extract data into a working dataframe. It only produces
hypotheses for the lead agent to confirm with the user before any math
runs (compute.py, owned by another component, takes a long-format CSV instead).

CLI:
    python3 survey.py <path-to-xlsx-or-csv> [--json]
    python3 survey.py <path> --self-test

Stdlib only, plus openpyxl for .xlsx reading.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import re
import sys
import time
from typing import Any

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError as _e:
    print(
        f"ERROR: openpyxl is required. Install with: pip install openpyxl  ({_e})",
        file=sys.stderr,
    )
    sys.exit(2)


# ---------------------------------------------------------------------------
# Constants & helpers
# ---------------------------------------------------------------------------

# Summary/total row labels. Matched against the WHOLE cell (with a small
# whitelist of trailing unit words) so a real customer named "Total Wine & More"
# or "Sum Ventures" is NOT mistaken for a total row — only cells that are
# essentially just the keyword (optionally + a unit) count.
_SUMMARY_TERMS = (
    "grand total", "subtotal", "sub-total", "total", "sum", "summary",
    "acv", "arr", "mrr", "average", "avg", "count",
)
_SUMMARY_SUFFIX = (
    r"(?:\s+(?:mrr|arr|acv|revenue|recurring|re-occurring|reoccurring|"
    r"customers?|accounts?|logos?|count|amount|rate|%|\$|usd))*"
)

CUSTOMER_HEADER_HINTS = [
    "customer id",
    "customer_id",
    "customerid",
    "customer",
    "client id",
    "client",
    "account id",
    "account",
    "id",
    "name",
]


def is_summary_label(value: Any) -> bool:
    """True only when the whole cell is a summary/total label (keyword + optional
    unit words and trailing punctuation), so customer names that merely START
    with a keyword don't false-match."""
    if not isinstance(value, str):
        return False
    s = value.strip().lower().rstrip(":").strip()
    if not s:
        return False
    for term in _SUMMARY_TERMS:
        if re.match(rf"^{re.escape(term)}{_SUMMARY_SUFFIX}\s*$", s):
            return True
    return False


# Rollforward / movement section labels that sometimes sit *inside* the customer
# column (above or below the customer rows) but are NOT customers. Matched
# position-agnostically — we never assume sections live at the top or bottom.
_MOVEMENT_TERMS = (
    "new", "net new", "upsell", "up-sell", "expansion", "expand",
    "downsell", "down-sell", "contraction", "contract", "churn", "churned",
    "gross churn", "net churn", "movement", "delta", "pipeline", "bookings",
    "beginning", "opening", "starting", "ending", "closing", "logo",
    "gross retention", "net retention", "grr", "nrr", "retention",
    "reactivation", "resurrection", "winback", "win-back", "reverse",
    "existing",
)
# Only allow a small set of trailing words so we match "New customers" / "Churn
# MRR" but NOT a real customer named "New Relic" or "Churnzero Inc".
_MOVEMENT_SUFFIX = r"(?:\s+(?:customers?|revenue|mrr|arr|acv|logos?|accounts?|rate|%|\$))*"


def is_movement_label(value: Any) -> bool:
    """True if the label is a rollforward/movement section header (anchored to
    the whole cell, so it won't fire on company names that merely start with a
    movement word)."""
    if not isinstance(value, str):
        return False
    s = value.strip().lower()
    if not s:
        return False
    for term in _MOVEMENT_TERMS:
        if re.match(rf"^{re.escape(term)}{_MOVEMENT_SUFFIX}\s*$", s):
            return True
    return False


def is_non_customer_label(value: Any) -> bool:
    """A row whose customer-column label is a summary OR a movement-section
    label is not a customer, regardless of where it sits on the sheet."""
    return is_summary_label(value) or is_movement_label(value)


def _row_has_monthly_data(ws, r: int, source_first_col: int, source_last_col: int) -> bool:
    """True if row r carries at least one numeric value across the source date
    columns. This is what separates a real customer row from a dataless label
    (a segment sub-header like 'Enterprise', a spacer, a stray note) that merely
    happens to have text in the customer column."""
    if not source_first_col or not source_last_col:
        return False
    for c in range(source_first_col, source_last_col + 1):
        v = ws.cell(row=r, column=c).value
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return True
    return False


def _detect_total_rows(ws, rows: list[int], source_first_col: int,
                       source_last_col: int, rel_tol: float = 0.005) -> list[int]:
    """Flag any row in `rows` whose per-month values equal the SUM of all the
    OTHER rows' values across the source months — i.e. a column-total row that
    slipped into the customer block without a label our vocabulary recognizes.

    Vocabulary-independent and conservative: requires at least 3 OTHER rows (so
    a single big customer that coincidentally equals the sum of two others is
    not mistaken for a total), the match to hold across every month the row has
    data (>=2 months), and exactly ONE row claiming to be the total of the rest
    (otherwise it's ambiguous — all-equal rows, nested subtotals — so we don't
    guess)."""
    if len(rows) < 4 or not source_first_col or not source_last_col:
        return []
    ncols = source_last_col - source_first_col + 1
    vals: dict[int, list[float]] = {}
    for r in rows:
        rv = []
        for c in range(source_first_col, source_last_col + 1):
            v = ws.cell(row=r, column=c).value
            rv.append(float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else 0.0)
        vals[r] = rv
    col_total = [sum(vals[r][j] for r in rows) for j in range(ncols)]

    totals: list[int] = []
    for r in rows:
        checked = 0
        matched = 0
        for j in range(ncols):
            mine = vals[r][j]
            others = col_total[j] - mine
            if abs(mine) < 1e-9 and abs(others) < 1e-9:
                continue
            checked += 1
            if abs(mine - others) <= rel_tol * max(abs(mine), 1.0):
                matched += 1
        if checked >= 2 and matched == checked:
            totals.append(r)
    return totals if len(totals) == 1 else []


def _split_runs_by_blanks(candidate_rows: list[int], blank_rows: set) -> list[list[int]]:
    """Split an ordered list of candidate customer rows into runs, breaking a run
    wherever a blank customer-column row sits between two consecutive candidates.
    Blank rows are the cleanest, label-independent block boundary: a top/bottom
    roll-up cluster is almost always separated from the real customer list by one
    or more blank rows."""
    runs: list[list[int]] = []
    current: list[int] = []
    for r in candidate_rows:
        if current and any(b in blank_rows for b in range(current[-1] + 1, r)):
            runs.append(current)
            current = []
        current.append(r)
    if current:
        runs.append(current)
    return runs


def classify_customer_rows(ws, header_row: int, customer_col: int | None,
                           source_first_col: int, source_last_col: int) -> tuple[list[int], list[dict], list[dict]]:
    """Split the rows below the header into real customer rows, section/summary
    rows, and excluded non-customer clusters.

    A row is a *candidate* customer only when its customer-column cell is
    non-empty, is NOT a summary/movement label, AND it carries monthly data (at
    least one numeric value across the source date columns) — the monthly-data
    requirement drops dataless label rows (segment sub-headers, spacers).

    Candidates are grouped into runs separated by blank rows. Every run that is
    comparably large (>= 25% of the biggest, or the biggest itself) is kept and
    merged into the customer block — so a stray/cosmetic blank inside a real
    customer list does NOT lop off half of it. Small runs that sit apart from the
    data (a roll-up/summary cluster, whose labels may be too verbose to
    pattern-match) are dropped and reported in excluded_clusters. Finally a
    structural total-row guard removes any kept row whose values equal the sum of
    the others (a column total that slipped in without a recognizable label).

    Returns (customer_rows, section_rows, excluded_clusters) where
    section_rows = [{row, label[, reason]}] and
    excluded_clusters = [{first_row, last_row, count}]."""
    customer_rows: list[int] = []
    section_rows: list[dict] = []
    excluded_clusters: list[dict] = []
    if customer_col is None:
        return customer_rows, section_rows, excluded_clusters

    candidates: list[int] = []
    blank_rows: set = set()
    for r in range(header_row + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=customer_col).value
        if v in (None, ""):
            blank_rows.add(r)
            continue
        if is_non_customer_label(v):
            section_rows.append({"row": r, "label": str(v).strip()})
            continue
        if not _row_has_monthly_data(ws, r, source_first_col, source_last_col):
            # Named, but no monthly data → a sub-header / segment label / spacer,
            # not a customer. Skip it (it anchors nothing).
            continue
        candidates.append(r)

    if not candidates:
        return customer_rows, section_rows, excluded_clusters

    runs = _split_runs_by_blanks(candidates, blank_rows)
    largest = max(len(r) for r in runs)
    # Keep every run that is comparably large (>= 25% of the biggest, min 3) OR is
    # itself the biggest. This merges a real customer list back together when a
    # stray/cosmetic blank row splits it, while still dropping a small summary
    # cluster that sits apart from the data. Small runs are reported, never
    # silently swallowed.
    keep_threshold = max(3, 0.25 * largest)
    kept_rows: list[int] = []
    for run in runs:
        if len(run) >= keep_threshold or len(run) == largest:
            kept_rows.extend(run)
        else:
            excluded_clusters.append(
                {"first_row": run[0], "last_row": run[-1], "count": len(run)}
            )
    kept_rows.sort()

    customer_rows = kept_rows
    # Structural total-row guard: drop a column-total row that slipped into the
    # block without a recognizable label (caught by summation, not vocabulary).
    total_rows = _detect_total_rows(ws, customer_rows, source_first_col, source_last_col)
    if total_rows:
        total_set = set(total_rows)
        customer_rows = [r for r in customer_rows if r not in total_set]
        for r in total_rows:
            lbl = ws.cell(row=r, column=customer_col).value
            section_rows.append({
                "row": r,
                "label": str(lbl).strip() if lbl not in (None, "") else "(unlabeled)",
                "reason": "values equal the sum of the other rows — looks like a column total",
            })
        section_rows.sort(key=lambda x: x["row"])
    return customer_rows, section_rows, excluded_clusters


def looks_like_date(value: Any) -> bool:
    """Return True if the cell value looks like a month-end date or month label."""
    if isinstance(value, dt.datetime) or isinstance(value, dt.date):
        return True
    if isinstance(value, str):
        s = value.strip()
        # e.g. Jan-25, January 2025, 2025-01, 2025-01-31, 1/2025, etc.
        patterns = [
            r"^[A-Za-z]{3,9}[-\s/]\d{2,4}$",          # Jan-25 / January 2025
            r"^\d{4}[-/]\d{1,2}([-/]\d{1,2})?$",      # 2025-01 / 2025-01-31
            r"^\d{1,2}[-/]\d{4}$",                    # 1/2025
            r"^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$",       # 1/31/2025
            # ISO datetime stringified (openpyxl sometimes returns these from
            # workbooks that stored dates as text, e.g. '2025-01-01 00:00:00')
            r"^\d{4}-\d{2}-\d{2}([T ]\d{2}:\d{2}(:\d{2})?(\.\d+)?Z?)?$",
            r"^\d{1,2}/\d{1,2}/\d{2,4}\s+\d{1,2}:\d{2}(:\d{2})?$",
        ]
        for p in patterns:
            if re.match(p, s):
                return True
    return False


def to_date(value: Any) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        s = value.strip()
        for fmt in (
            "%Y-%m-%d %H:%M:%S",  # ISO datetime stringified
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d",
            "%Y-%m",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y",
            "%m-%d-%Y",
            "%b-%y",
            "%b %Y",
            "%B %Y",
            "%b-%Y",
        ):
            try:
                return dt.datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None


def fmt_month(d: dt.date) -> str:
    """Format date as e.g. 'Jan-25'."""
    return d.strftime("%b-%y")


# ---------------------------------------------------------------------------
# Sheet inspection
# ---------------------------------------------------------------------------

# Sheet-name hints used by the cheap Pass 1 scorer.
SOURCE_NAME_HINTS = (
    "raw", "data", "customer", "revenue", "mrr", "arr", "billing",
    "subscription", "invoice", "transaction",
)
NON_SOURCE_NAME_HINTS = (
    "summary", "metric", "rollup", "rollforward", "corkscrew", "dashboard",
    "calc", "scratch", "notes", "instruction", "cover", "readme", "config",
    "chart", "graph", "pivot",
)


def quick_score_sheet(ws) -> dict:
    """Cheap structural scan — does NOT do full inspection.

    Looks for a date-like header row, counts data rows beneath it, and applies
    a sheet-name bonus/penalty. Higher score = more likely to be the source
    customer × month matrix.

    This is Pass 1 of the two-pass survey: every sheet gets scored cheaply so
    we can pick the top candidate(s) for deep inspection.
    """
    header_row, date_run = find_date_header_row(ws)
    name = ws.title
    name_lower = name.lower()

    if not date_run:
        return {
            "name": name,
            "score": 0.0,
            "date_count": 0,
            "n_data_rows": 0,
            "header_row": 0,
            "reason": "No date-like header row.",
        }

    date_count = len(date_run)
    n_data_rows = max(0, ws.max_row - header_row)

    # Name bonus: source-y names lift, summary-y names penalise.
    name_bonus = 0.0
    if any(h in name_lower for h in SOURCE_NAME_HINTS):
        name_bonus += 10.0
    if any(h in name_lower for h in NON_SOURCE_NAME_HINTS):
        name_bonus -= 15.0

    # Rows factor capped — diminishing returns past ~50 rows.
    rows_factor = min(n_data_rows / 5.0, 10.0)

    score = float(date_count) * rows_factor + name_bonus

    return {
        "name": name,
        "score": round(score, 2),
        "date_count": date_count,
        "n_data_rows": n_data_rows,
        "header_row": header_row,
        "reason": (
            f"{date_count} date col(s), {n_data_rows} rows beneath header, "
            f"name bonus {name_bonus:+.0f}"
        ),
    }


def find_date_header_row(ws) -> tuple[int, list[tuple[int, dt.date]]]:
    """
    Scan the first 30 rows. Find the row that has the longest contiguous run of
    date-like cells. Return (row_number, list of (col_index, date) tuples for
    the leftmost contiguous run).
    """
    best_row = None
    best_run: list[tuple[int, dt.date]] = []

    max_scan_rows = min(30, ws.max_row)
    max_scan_cols = ws.max_column

    for r in range(1, max_scan_rows + 1):
        # Collect contiguous runs of date cells in this row
        runs: list[list[tuple[int, dt.date]]] = []
        current: list[tuple[int, dt.date]] = []
        for c in range(1, max_scan_cols + 1):
            v = ws.cell(row=r, column=c).value
            d = to_date(v) if (looks_like_date(v) or isinstance(v, (dt.date, dt.datetime))) else None
            if d is not None:
                current.append((c, d))
            else:
                if current:
                    runs.append(current)
                    current = []
        if current:
            runs.append(current)
        # Pick the longest run for this row
        if runs:
            longest = max(runs, key=len)
            if len(longest) >= 2 and len(longest) > len(best_run):
                # Prefer the leftmost run when this row also has multiple equal-length runs
                # (longest is already the max; for ties Python's max keeps the first occurrence)
                best_row = r
                best_run = longest
    return best_row or 0, best_run


def find_customer_column(ws, header_row: int, first_date_col: int) -> dict | None:
    """
    Look for a customer ID column in the rows immediately above the header_row
    (typical: header is one row above first data row), and to the left of the
    first date column. Heuristics:
      1. Header text matches CUSTOMER_HEADER_HINTS (in header_row or row above).
      2. Otherwise, leftmost non-empty column to the left of first_date_col with
         non-empty values below header_row.
    """
    candidate = None

    # Look at header_row and the row above
    for hr in (header_row, header_row - 1):
        if hr < 1:
            continue
        for c in range(1, max(first_date_col, 2)):
            v = ws.cell(row=hr, column=c).value
            if isinstance(v, str):
                s = v.strip().lower()
                for hint in CUSTOMER_HEADER_HINTS:
                    if hint in s:
                        return {
                            "col_letter": get_column_letter(c),
                            "col_index": c,
                            "evidence": f"Header at {get_column_letter(c)}{hr}: {v!r}",
                        }

    # Fallback: leftmost column with non-empty data below header_row
    data_start = header_row + 1
    data_end = ws.max_row
    for c in range(1, max(first_date_col, 2)):
        non_empty = 0
        for r in range(data_start, min(data_end, data_start + 50) + 1):
            if ws.cell(row=r, column=c).value not in (None, ""):
                non_empty += 1
        if non_empty >= 3:
            candidate = {
                "col_letter": get_column_letter(c),
                "col_index": c,
                "evidence": (
                    f"Leftmost column with non-empty data below row {header_row} "
                    f"(no explicit header label found)"
                ),
            }
            break
    return candidate


def detect_summary_rows(ws, customer_col: int | None) -> list[dict]:
    """Find rows whose first-or-customer-column label looks like a total/summary label."""
    out = []
    label_col = customer_col if customer_col else 1
    for r in range(1, ws.max_row + 1):
        # check label_col first, then column A as a fallback
        for c in (label_col, 1):
            v = ws.cell(row=r, column=c).value
            if is_summary_label(v):
                out.append({"row": r, "label": str(v).strip()})
                break
    return out


def detect_derived_blocks(ws, header_row: int, source_first_col: int, source_last_col: int) -> list[dict]:
    """
    Identify derived (computed) blocks of repeated date columns to the right of
    the source matrix. The label is taken from header_row-1 (typical sectioned-
    wide layout). Empty labels are reported as 'None' (the spec calls this out:
    the leftmost block may have an empty header label even though the columns
    are real).

    Strategy: walk left-to-right starting at column 1; group contiguous date-
    columns into blocks; for each block read its label from row above the
    header. The first block is the source; everything else is derived.
    """
    label_row = header_row - 1 if header_row > 1 else None

    # Build contiguous run list across header_row
    runs: list[tuple[int, int]] = []
    current_start = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if to_date(v) is not None or looks_like_date(v):
            if current_start is None:
                current_start = c
        else:
            if current_start is not None:
                runs.append((current_start, c - 1))
                current_start = None
    if current_start is not None:
        runs.append((current_start, ws.max_column))

    blocks = []
    for (start, end) in runs:
        # Pull the most common non-empty label across the block at label_row
        label = None
        if label_row:
            labels_seen = []
            for c in range(start, end + 1):
                lv = ws.cell(row=label_row, column=c).value
                if isinstance(lv, str) and lv.strip():
                    labels_seen.append(lv.strip())
            if labels_seen:
                # Use the modal value
                label = max(set(labels_seen), key=labels_seen.count)
        is_source = (start == source_first_col and end == source_last_col)
        blocks.append({
            "col_range": f"{get_column_letter(start)}-{get_column_letter(end)}",
            "first_col": start,
            "last_col": end,
            "label": label if label else ("" if is_source else "None"),
            "type": "source" if is_source else "derived/computed",
        })
    return blocks


def detect_negatives(ws, header_row: int, customer_col: int | None,
                     source_first_col: int, source_last_col: int,
                     customer_row_set: set) -> list[dict]:
    """Walk the source block and flag negative numeric values, tagging each by
    whether it falls in a real customer row vs a section/summary row. Negatives
    in section rows (e.g. a "Churn" or "Contraction" line) are expected and
    should not force a user decision — only customer-row negatives matter."""
    out = []
    data_start = header_row + 1
    for r in range(data_start, ws.max_row + 1):
        if customer_col is not None:
            cust_val = ws.cell(row=r, column=customer_col).value
            if cust_val in (None, ""):
                continue
        else:
            cust_val = None
        for c in range(source_first_col, source_last_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, (int, float)) and not isinstance(v, bool) and v < 0:
                out.append({
                    "row": r,
                    "col": c,
                    "col_letter": get_column_letter(c),
                    "value": float(v),
                    "customer": cust_val,
                    "in_customer_row": r in customer_row_set,
                })
    return out


def assess_scale(ws, header_row: int, source_first_col: int,
                 source_last_col: int, summary_rows: list[dict]) -> dict:
    """
    Return a verdict on whether the source values are MRR or ARR.
    Strategy:
      1. If a 'Total MRR' AND 'Total ARR' summary row both exist with a numeric
         ratio ~12, the leftmost-block scale matches whichever the rows are
         pulling from. Inspect first-column values: if they match Total MRR
         column for the same date, the data is MRR.
      2. Otherwise fall back to magnitude heuristic: typical SaaS customer-
         month MRR is hundreds-to-tens-of-thousands; ARR is twelve times that.
    """
    total_mrr_row = None
    total_arr_row = None
    for sr in summary_rows:
        lab = sr["label"].lower()
        if "total mrr" in lab or lab == "mrr":
            total_mrr_row = sr["row"]
        elif "total arr" in lab or lab == "arr":
            total_arr_row = sr["row"]

    if total_mrr_row and total_arr_row:
        c = source_first_col
        mrr_val = ws.cell(row=total_mrr_row, column=c).value
        arr_val = ws.cell(row=total_arr_row, column=c).value
        if isinstance(mrr_val, (int, float)) and isinstance(arr_val, (int, float)) and mrr_val:
            ratio = arr_val / mrr_val
            if 11.5 <= ratio <= 12.5:
                # The leftmost block matches the Total MRR row → MRR
                # (Confirm by summing column for non-summary rows and comparing.)
                col_sum = 0.0
                for r in range(header_row + 1, ws.max_row + 1):
                    if r in {total_mrr_row, total_arr_row}:
                        continue
                    v = ws.cell(row=r, column=c).value
                    if isinstance(v, (int, float)):
                        col_sum += float(v)
                # Compare to Total MRR vs Total ARR
                d_mrr = abs(col_sum - mrr_val)
                d_arr = abs(col_sum - arr_val)
                if d_mrr < d_arr:
                    return {
                        "verdict": "MRR",
                        "evidence": (
                            f"Total MRR row at R{total_mrr_row}, Total ARR row at R{total_arr_row}, "
                            f"Total ARR / Total MRR ratio = {ratio:.2f}; column {get_column_letter(c)} "
                            f"sum {col_sum:,.2f} matches Total MRR ({mrr_val:,.2f})"
                        ),
                    }
                else:
                    return {
                        "verdict": "ARR",
                        "evidence": (
                            f"Total MRR row at R{total_mrr_row}, Total ARR row at R{total_arr_row}, "
                            f"ratio={ratio:.2f}; column {get_column_letter(c)} sum {col_sum:,.2f} "
                            f"matches Total ARR ({arr_val:,.2f})"
                        ),
                    }

    # Fallback: magnitude check.
    samples = []
    for r in range(header_row + 1, min(ws.max_row, header_row + 30) + 1):
        v = ws.cell(row=r, column=source_first_col).value
        if isinstance(v, (int, float)) and v > 0:
            samples.append(float(v))
    if samples:
        median = sorted(samples)[len(samples) // 2]
        if median < 25_000:
            return {
                "verdict": "MRR (hypothesis)",
                "evidence": (
                    f"No Total MRR/ARR rows found. Median per-customer first-month value is "
                    f"{median:,.0f}, which fits MRR scale (under $25K/mo)."
                ),
            }
        else:
            return {
                "verdict": "ARR (hypothesis)",
                "evidence": (
                    f"No Total MRR/ARR rows found. Median per-customer first-month value is "
                    f"{median:,.0f}, which fits ARR scale (over $25K)."
                ),
            }
    return {"verdict": "unknown", "evidence": "Could not determine scale from headers, totals, or magnitude."}


# ---------------------------------------------------------------------------
# Sheet shape & role hypothesis
# ---------------------------------------------------------------------------

def hypothesize_shape(ws, header_row: int, source_first_col: int,
                      source_last_col: int, blocks: list[dict],
                      sheet_name: str) -> tuple[str, str, str]:
    """
    Returns (role, shape, confidence).
      role  ∈ {source data, finished analysis, lookup, scratch, irrelevant}
      shape ∈ {single-table, sectioned-wide, pivoted, pre-aggregated-rollforward}
    """
    name_lower = sheet_name.lower()

    # Check if sheet looks like a corkscrew / rollforward (row labels include
    # Beginning, Ending, Churn, etc., in column A or B, with dates across).
    rollforward_terms = ("beginning", "ending", "new arr", "expansion", "contraction", "churn", "rollforward", "corkscrew")
    label_hits = 0
    for r in range(1, min(ws.max_row, 30) + 1):
        for c in (1, 2):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                vl = v.lower()
                for term in rollforward_terms:
                    if term in vl:
                        label_hits += 1
                        break
    if label_hits >= 3 or "corkscrew" in name_lower or "rollforward" in name_lower:
        return ("finished analysis", "pre-aggregated rollforward", "high")

    # Sectioned-wide if multiple horizontal date-blocks present.
    if len(blocks) >= 2:
        return ("source data", "sectioned-wide", "high")

    # Single-table source data if exactly one date block + customer column found.
    if len(blocks) == 1 and source_first_col and source_last_col:
        return ("source data", "single-table", "high")

    # Otherwise scratch / irrelevant.
    return ("irrelevant", "single-table", "low")


# ---------------------------------------------------------------------------
# Per-sheet inspection
# ---------------------------------------------------------------------------

def inspect_sheet(ws, as_of: dt.date | None = None) -> dict:
    as_of = as_of or dt.date.today()
    name = ws.title
    rows = ws.max_row
    cols = ws.max_column

    # Phase 1: shape + role hypothesis ---------------------------------------
    header_row, date_run = find_date_header_row(ws)
    if not date_run:
        # No date columns at all
        return {
            "name": name,
            "dimensions": {"rows": rows, "cols": cols},
            "hypothesis": {
                "role": "irrelevant",
                "shape": "single-table",
                "confidence": "low",
            },
            "candidate_customer_column": None,
            "candidate_date_columns": None,
            "scale_signal": {"verdict": "n/a", "evidence": "No date-like header row detected"},
            "summary_rows_detected": [],
            "derived_blocks_detected": [],
            "negative_values": [],
            "sufficiency": {
                "verdict": "fail",
                "reason": "No identifiable date columns — sheet is not a customer × month matrix.",
            },
        }

    # The leftmost contiguous run of date cells in header_row is the source.
    source_first_col = date_run[0][0]
    source_last_col = date_run[-1][0]
    first_date = date_run[0][1]
    last_date = date_run[-1][1]
    date_count = len(date_run)

    # Detect blocks across the full sheet (for sectioned-wide layouts).
    blocks = detect_derived_blocks(ws, header_row, source_first_col, source_last_col)

    # Customer column.
    customer_col_info = find_customer_column(ws, header_row, source_first_col)
    customer_col_index = customer_col_info["col_index"] if customer_col_info else None

    # Summary rows (within source block).
    summary_rows = detect_summary_rows(ws, customer_col_index)

    # Position-agnostic split of rows below the header into real customers vs
    # section/summary rows (works whether sections sit above, below, or among
    # the customer rows — never assumes a fixed position).
    customer_rows, section_rows_in_col, excluded_clusters = classify_customer_rows(
        ws, header_row, customer_col_index, source_first_col, source_last_col
    )
    customer_row_set = set(customer_rows)

    # Negative values, tagged customer-row vs section-row.
    negatives = detect_negatives(
        ws, header_row, customer_col_index, source_first_col, source_last_col, customer_row_set
    )
    customer_negatives = [n for n in negatives if n.get("in_customer_row")]

    # Scale verdict.
    scale = assess_scale(ws, header_row, source_first_col, source_last_col, summary_rows)

    # Role + shape hypothesis.
    role, shape, confidence = hypothesize_shape(
        ws, header_row, source_first_col, source_last_col, blocks, name
    )

    # Customer count + contiguous range, excluding section rows. Only meaningful
    # when this looks like source data.
    customer_count = None
    customer_row_range = None
    if role == "source data" and customer_col_index:
        customer_count = len(customer_rows)
        if customer_rows:
            first_r, last_r = customer_rows[0], customer_rows[-1]
            customer_row_range = {
                "first_row": first_r,
                "last_row": last_r,
                "contiguous": (last_r - first_r + 1) == len(customer_rows),
                "section_rows_excluded": len(section_rows_in_col),
            }

    # Forecast / incompleteness detection. The current calendar month is still
    # IN PROGRESS, so it is NOT a complete booked actual — the last complete
    # month is the one *before* as_of's month (e.g. as_of Jun-26 → last complete
    # is May-26). Any column in the current month or later (genuine future) is
    # flagged as not-a-complete-actual. `as_of` is injectable so this stays
    # deterministic under test; production defaults to today.
    def _is_incomplete(d: dt.date) -> bool:
        return (d.year, d.month) >= (as_of.year, as_of.month)

    future_cols = [
        {
            "col_letter": get_column_letter(c),
            "month": fmt_month(d),
            "kind": (
                "current-month-in-progress"
                if (d.year, d.month) == (as_of.year, as_of.month)
                else "future"
            ),
        }
        for (c, d) in date_run if _is_incomplete(d)
    ]
    complete_dates = [d for (c, d) in date_run if not _is_incomplete(d)]
    forecast = {
        "as_of": as_of.isoformat(),
        "future_columns": future_cols,
        "actuals_through": fmt_month(max(complete_dates)) if complete_dates else None,
    }

    # Per-sheet sufficiency verdict (only customer-row negatives gate it).
    sheet_suff = sheet_sufficiency(role, customer_count, date_count,
                                   customer_negatives, future_cols)

    # Derived-block list excludes the source block.
    derived_blocks = [b for b in blocks if b["type"] != "source"]

    return {
        "name": name,
        "dimensions": {"rows": rows, "cols": cols},
        "hypothesis": {"role": role, "shape": shape, "confidence": confidence},
        "candidate_customer_column": customer_col_info,
        "candidate_date_columns": (
            {
                "first_col": source_first_col,
                "last_col": source_last_col,
                "first_col_letter": get_column_letter(source_first_col),
                "last_col_letter": get_column_letter(source_last_col),
                "count": date_count,
                "range": f"{fmt_month(first_date)} to {fmt_month(last_date)}",
                "header_row": header_row,
            }
            if date_run else None
        ),
        "scale_signal": scale,
        "summary_rows_detected": summary_rows,
        "derived_blocks_detected": derived_blocks,
        "negative_values": negatives,
        "customer_count_hypothesis": customer_count,
        "customer_row_range": customer_row_range,
        "section_rows_in_customer_col": section_rows_in_col,
        "excluded_row_clusters": excluded_clusters,
        "forecast": forecast,
        "sufficiency": sheet_suff,
    }


def sheet_sufficiency(role: str, customer_count: int | None,
                      date_count: int, customer_negatives: list[dict],
                      future_cols: list[dict] | None = None) -> dict:
    """Per-sheet sufficiency for retention analysis. Only customer-row negatives
    are surfaced; section-row negatives (e.g. a 'Churn' line) are expected."""
    future_cols = future_cols or []
    if role != "source data":
        return {
            "verdict": "n/a",
            "reason": f"Sheet role hypothesized as '{role}' — not the analysis input.",
        }
    if customer_count is None or customer_count < 1:
        return {"verdict": "fail", "reason": "No customer rows detected."}
    if date_count < 2:
        return {"verdict": "fail", "reason": f"Only {date_count} date column — need at least 2 distinct months."}
    caveats = []
    if date_count < 12:
        caveats.append(
            f"Only {date_count} months — LTM (last twelve months) retention will not be available; "
            "monthly-only with the limitation called out."
        )
    if customer_negatives:
        caveats.append(f"{len(customer_negatives)} negative value(s) in customer rows — discuss treatment with user.")
    if future_cols:
        caveats.append(
            f"{len(future_cols)} month column(s) are not complete actuals "
            f"({future_cols[0]['month']} onward — the current month is still in progress and/or "
            "later months are forecast). Confirm the actuals cutoff with the user."
        )
    if caveats:
        return {"verdict": "pass-with-caveats", "reason": " ".join(caveats)}
    return {
        "verdict": "pass",
        "reason": (
            f"{customer_count} customers × {date_count} months of revenue detected — "
            f"sufficient for monthly{' and LTM' if date_count >= 13 else ''} retention."
        ),
    }


# ---------------------------------------------------------------------------
# CSV inspection
# ---------------------------------------------------------------------------

def inspect_csv(path: str, as_of: dt.date | None = None) -> dict:
    """
    CSV path handling. We support two CSV shapes:
      1. Wide (customer per row, date columns)  → wrap into a one-sheet workbook view.
      2. Long (customer_id, month, mrr)          → flagged as already-normalized input.
    """
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    if not rows:
        return {
            "file": path,
            "sheets": [],
            "overall_sufficiency": {"verdict": "fail", "reason": "CSV is empty."},
        }

    header = rows[0]
    header_lower = [str(h).strip().lower() for h in header]
    n_rows = len(rows)
    n_cols = len(header)

    # Detect long-format
    is_long = (
        len(header) >= 3
        and any("customer" in h or h == "id" for h in header_lower)
        and any(h in {"month", "date", "period"} for h in header_lower)
        and any(h in {"mrr", "arr", "revenue", "amount"} for h in header_lower)
    )

    if is_long:
        sheet = {
            "name": os.path.basename(path),
            "dimensions": {"rows": n_rows, "cols": n_cols},
            "hypothesis": {"role": "source data", "shape": "long-format", "confidence": "high"},
            "candidate_customer_column": {
                "col_letter": "A", "col_index": 1,
                "evidence": f"Long-format CSV header: {header}",
            },
            "candidate_date_columns": None,
            "scale_signal": {"verdict": "MRR (hypothesis)", "evidence": "Header named 'mrr' (column-name signal)"},
            "summary_rows_detected": [],
            "derived_blocks_detected": [],
            "negative_values": [],
            "customer_count_hypothesis": None,
            "sufficiency": {
                "verdict": "pass",
                "reason": "Long-format CSV — customer_id/month/mrr columns detected. Ready for compute.py.",
            },
        }
        return {
            "file": path,
            "sheets": [sheet],
            "overall_sufficiency": sheet["sufficiency"],
        }

    # Wide-format CSV: write to a temporary workbook view in memory.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = os.path.basename(path)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            # Try to interpret numeric cells
            try:
                if val == "" or val is None:
                    cast: Any = None
                else:
                    s = str(val).strip().replace(",", "").replace("$", "")
                    if re.match(r"^-?\d+(\.\d+)?$", s):
                        cast = float(s)
                    else:
                        cast = val
            except Exception:
                cast = val
            ws.cell(row=r_idx, column=c_idx).value = cast
    sheet_report = inspect_sheet(ws, as_of=as_of)
    return {
        "file": path,
        "sheets": [sheet_report],
        "overall_sufficiency": sheet_report["sufficiency"],
    }


# ---------------------------------------------------------------------------
# Top-level inspect
# ---------------------------------------------------------------------------

def inspect_file(path: str, as_of: dt.date | None = None) -> dict:
    """Two-pass survey: cheap score every sheet, then deep-inspect only the
    top candidate (plus any close runner-up). Skips sheets that clearly aren't
    the source customer × month matrix so we don't waste time profiling
    pre-built dashboards, instructions, etc.

    `as_of` is injectable (defaults to today) so the forecast/current-month
    detection is deterministic under test.
    """
    if path.lower().endswith(".csv"):
        return inspect_csv(path, as_of=as_of)
    if not path.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError(f"Unsupported file extension: {path}. Provide .xlsx or .csv.")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)

    # Pass 1: cheap structural scan of every sheet.
    quick_scores = [quick_score_sheet(wb[name]) for name in wb.sheetnames]
    quick_scores.sort(key=lambda q: q["score"], reverse=True)

    # Decide which sheets get the expensive Pass 2 deep inspect.
    # Rule: top scorer always. Plus any runner-up scoring >= 80% of the top —
    # that's a "near tie" and the user should see both so they can override.
    NEAR_TIE_RATIO = 0.8
    if not quick_scores or quick_scores[0]["score"] <= 0:
        # No sheet looks like source data — deep-inspect the top one anyway so
        # the user gets a useful failure report.
        deep_names = [quick_scores[0]["name"]] if quick_scores else []
        near_tie = False
    else:
        top_score = quick_scores[0]["score"]
        threshold = NEAR_TIE_RATIO * top_score
        deep_names = [q["name"] for q in quick_scores
                      if q["score"] >= threshold and q["score"] > 0]
        deep_names = deep_names[:2]  # cap at 2 — three-way tie is contrived
        near_tie = len(deep_names) > 1

    # Pass 2: deep inspect the chosen sheet(s) only.
    sheet_reports = [inspect_sheet(wb[name], as_of=as_of) for name in deep_names]

    # Skipped sheets: everything we didn't deep-inspect, with their Pass-1
    # signals so the model can show them to the user when relevant.
    skipped_sheets = [
        {
            "name": q["name"],
            "score": q["score"],
            "reason": q["reason"],
        }
        for q in quick_scores
        if q["name"] not in deep_names
    ]

    overall = overall_sufficiency(sheet_reports)
    return {
        "file": path,
        "sheets": sheet_reports,
        "skipped_sheets": skipped_sheets,
        "near_tie": near_tie,
        "overall_sufficiency": overall,
    }


def overall_sufficiency(sheets: list[dict]) -> dict:
    """If any source-data sheet passes, overall passes (using the best one)."""
    best = None
    for s in sheets:
        if s["hypothesis"]["role"] != "source data":
            continue
        v = s["sufficiency"]["verdict"]
        rank = {"pass": 3, "pass-with-caveats": 2, "fail": 1, "n/a": 0}
        if best is None or rank.get(v, 0) > rank.get(best["sufficiency"]["verdict"], 0):
            best = s
    if best is None:
        return {
            "verdict": "fail",
            "reason": "No sheet hypothesized as source data — cannot run retention analysis.",
        }
    return {
        "verdict": best["sufficiency"]["verdict"],
        "reason": f"Best source-data sheet: {best['name']!r} — {best['sufficiency']['reason']}",
    }


# ---------------------------------------------------------------------------
# Pretty-print
# ---------------------------------------------------------------------------

def render_text(report: dict) -> str:
    lines = []
    lines.append("=" * 78)
    lines.append(f"FILE: {report['file']}")
    lines.append("=" * 78)
    lines.append("")
    lines.append("NOTE: All findings below are HYPOTHESES for the user to confirm at the")
    lines.append("Phase 1 / 4 checkpoints. Nothing is extracted into a working dataset here.")
    lines.append("")
    for s in report["sheets"]:
        lines.append("-" * 78)
        lines.append(f"SHEET: {s['name']}   ({s['dimensions']['rows']} rows × {s['dimensions']['cols']} cols)")
        lines.append("-" * 78)
        h = s["hypothesis"]
        lines.append(
            f"  Hypothesis (role / shape / confidence): "
            f"{h['role']} / {h['shape']} / {h['confidence']}"
        )
        if s.get("candidate_customer_column"):
            cc = s["candidate_customer_column"]
            lines.append(
                f"  Candidate customer column (HYP): {cc['col_letter']} (idx {cc['col_index']}) — {cc['evidence']}"
            )
        else:
            lines.append("  Candidate customer column: not detected")

        if s.get("candidate_date_columns"):
            dc = s["candidate_date_columns"]
            lines.append(
                f"  Candidate date columns (HYP): cols {dc['first_col_letter']}-{dc['last_col_letter']} "
                f"({dc['count']} months, {dc['range']}); header row {dc['header_row']}"
            )
        else:
            lines.append("  Candidate date columns: not detected")

        if s.get("customer_count_hypothesis") is not None:
            lines.append(f"  Customer-row count (HYP): {s['customer_count_hypothesis']}")
        crr = s.get("customer_row_range")
        if crr:
            note = "contiguous" if crr["contiguous"] else "NON-contiguous"
            extra = (f", {crr['section_rows_excluded']} section row(s) excluded"
                     if crr["section_rows_excluded"] else "")
            lines.append(
                f"  Customer block (HYP): rows {crr['first_row']}-{crr['last_row']} "
                f"({note}{extra})"
            )
        sect = s.get("section_rows_in_customer_col") or []
        if sect:
            labels = ", ".join(f"row {x['row']}:{x['label']!r}" for x in sect[:6])
            lines.append(f"  Section rows inside the customer column (NOT customers): {labels}")
        excl = s.get("excluded_row_clusters") or []
        if excl:
            clusters = ", ".join(
                f"rows {x['first_row']}-{x['last_row']} ({x['count']})" for x in excl
            )
            lines.append(
                f"  Excluded non-customer cluster(s) — separated from the customer "
                f"block by blank rows (likely a summary/roll-up block): {clusters}"
            )

        fc = s.get("forecast") or {}
        if fc.get("future_columns"):
            cols = ", ".join(x["month"] for x in fc["future_columns"])
            lines.append(
                f"  Incomplete/forecast tail (HYP): {len(fc['future_columns'])} month(s) are not "
                f"complete actuals ({cols}; current month in progress and/or later months forecast) "
                f"— last complete actual appears to be {fc.get('actuals_through')}. Confirm cutoff."
            )

        scale = s["scale_signal"]
        lines.append(f"  Scale signal (HYP): {scale['verdict']} — {scale['evidence']}")

        if s["summary_rows_detected"]:
            lines.append(f"  Summary rows detected ({len(s['summary_rows_detected'])}):")
            for sr in s["summary_rows_detected"]:
                lines.append(f"    - row {sr['row']}: {sr['label']!r}")
        else:
            lines.append("  Summary rows detected: none")

        if s["derived_blocks_detected"]:
            lines.append(f"  Derived blocks detected ({len(s['derived_blocks_detected'])}) — to EXCLUDE from analysis input:")
            for b in s["derived_blocks_detected"]:
                lines.append(f"    - cols {b['col_range']}: label={b['label']!r}, type={b['type']}")
        else:
            lines.append("  Derived blocks detected: none")

        negs = s["negative_values"]
        cust_negs = [n for n in negs if n.get("in_customer_row")]
        sect_negs = [n for n in negs if not n.get("in_customer_row")]
        if cust_negs:
            lines.append(f"  Negatives in CUSTOMER rows ({len(cust_negs)}) — need a user decision:")
            for n in cust_negs[:10]:
                lines.append(
                    f"    - {n['col_letter']}{n['row']} = {n['value']:,.2f} (customer={n['customer']!r})"
                )
            if len(cust_negs) > 10:
                lines.append(f"    ... and {len(cust_negs) - 10} more")
        else:
            lines.append("  Negatives in customer rows: none")
        if sect_negs:
            lines.append(
                f"  Negatives in section/summary rows ({len(sect_negs)}): expected "
                "(e.g. a Churn/Contraction line) — no user decision needed."
            )

        suf = s["sufficiency"]
        lines.append(f"  Sheet sufficiency: {suf['verdict'].upper()} — {suf['reason']}")
        lines.append("")

    # Skipped sheets — only surface on near-tie. Otherwise the user doesn't
    # need to see the noise.
    skipped = report.get("skipped_sheets", [])
    near_tie = report.get("near_tie", False)
    if skipped and near_tie:
        lines.append("-" * 78)
        lines.append(
            "NEAR-TIE: another sheet scored similarly to the chosen source — "
            "review both above and confirm."
        )
        lines.append("-" * 78)
        for sk in skipped:
            lines.append(f"  - {sk['name']!r} (score {sk['score']}) — {sk['reason']}")
        lines.append("")

    lines.append("=" * 78)
    o = report["overall_sufficiency"]
    lines.append(f"OVERALL SUFFICIENCY: {o['verdict'].upper()}")
    lines.append(f"  Reason: {o['reason']}")
    lines.append("=" * 78)
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Self-test
# ---------------------------------------------------------------------------

# Path to the bundled synthetic test fixture. Resolved relative to this file so
# it works whether the skill runs from an agent's skills folder or a fresh clone.
DEFAULT_FIXTURE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "fixtures",
    "sample_retention_data.xlsx",
)


def run_self_test(path: str | None = None) -> int:
    target = path or DEFAULT_FIXTURE_PATH
    print(f"Self-test target: {target}")
    print()
    if not os.path.exists(target):
        print(f"FAIL — test target not found: {target}")
        return 1

    # Pin as_of to just after the fixture's last month (Jun-26) so all 18 months
    # read as complete actuals — keeps the fixture assertions deterministic
    # regardless of the real calendar date the test runs on.
    report = inspect_file(target, as_of=dt.date(2026, 7, 15))
    print(render_text(report))
    print()

    failures: list[str] = []

    sheets_by_name = {s["name"]: s for s in report["sheets"]}

    # --- Raw Data assertions (per fixtures/EXPECTED_VALUES.md §5) ----------
    raw = sheets_by_name.get("Raw Data")
    if not raw:
        failures.append("Sheet 'Raw Data' not found.")
    else:
        h = raw["hypothesis"]
        if h["role"] != "source data":
            failures.append(f"Raw Data: role expected 'source data', got {h['role']!r}")
        if h["shape"] != "sectioned-wide":
            failures.append(f"Raw Data: shape expected 'sectioned-wide', got {h['shape']!r}")
        if h.get("confidence") != "high":
            failures.append(f"Raw Data: confidence expected 'high', got {h.get('confidence')!r}")
        cc = raw["candidate_customer_column"]
        if not cc or cc["col_letter"] != "B":
            failures.append(f"Raw Data: customer col expected 'B', got {cc!r}")
        if cc and cc.get("col_index") != 2:
            failures.append(f"Raw Data: customer col_index expected 2, got {cc.get('col_index')!r}")
        dc = raw["candidate_date_columns"]
        if not dc:
            failures.append("Raw Data: date columns not detected.")
        else:
            if dc["count"] != 18:
                failures.append(f"Raw Data: expected 18 date columns, got {dc['count']}")
            if dc["range"] != "Jan-25 to Jun-26":
                failures.append(
                    f"Raw Data: expected range 'Jan-25 to Jun-26', got {dc['range']!r}"
                )
            if dc["first_col_letter"] != "C" or dc["last_col_letter"] != "T":
                failures.append(
                    f"Raw Data: expected source date cols C-T, got "
                    f"{dc['first_col_letter']}-{dc['last_col_letter']}"
                )
            if dc.get("first_col") != 3 or dc.get("last_col") != 20:
                failures.append(
                    f"Raw Data: expected first_col=3 last_col=20, got "
                    f"{dc.get('first_col')}..{dc.get('last_col')}"
                )
            if dc.get("header_row") != 7:
                failures.append(f"Raw Data: header_row expected 7, got {dc.get('header_row')!r}")
        sc = raw["scale_signal"]
        # Accept both "MRR" and "MRR (hypothesis)" — the fixture currently
        # produces the parenthetical form because no Total-MRR/Total-ARR
        # corroboration rows exist in Raw Data.
        if not str(sc["verdict"]).startswith("MRR"):
            failures.append(
                f"Raw Data: scale verdict expected to start with 'MRR', got {sc['verdict']!r}"
            )
        # --- Derived blocks: exactly 5, in order, with the expected labels and col_ranges
        expected_blocks = [
            ("New customer MRR", "V-Y"),
            ("Upsell MRR", "AA-AD"),
            ("Downsell MRR", "AF-AI"),
            ("Churn MRR", "AK-AN"),
            ("Check", "AP-AS"),
        ]
        derived = raw["derived_blocks_detected"]
        if len(derived) != len(expected_blocks):
            failures.append(
                f"Raw Data: expected exactly {len(expected_blocks)} derived blocks, "
                f"got {len(derived)}"
            )
        else:
            for i, (exp_label, exp_range) in enumerate(expected_blocks):
                got = derived[i]
                if got.get("label") != exp_label:
                    failures.append(
                        f"Raw Data: derived_blocks_detected[{i}].label expected "
                        f"{exp_label!r}, got {got.get('label')!r}"
                    )
                if got.get("col_range") != exp_range:
                    failures.append(
                        f"Raw Data: derived_blocks_detected[{i}].col_range expected "
                        f"{exp_range!r}, got {got.get('col_range')!r}"
                    )
        # No negative values in the fixture.
        if raw.get("negative_values") != []:
            failures.append(
                f"Raw Data: expected negative_values == [], got {raw.get('negative_values')!r}"
            )
        # Customer-count hypothesis (10 IDs in B8:B17).
        if raw.get("customer_count_hypothesis") != 10:
            failures.append(
                f"Raw Data: customer_count_hypothesis expected 10, got "
                f"{raw.get('customer_count_hypothesis')!r}"
            )
        # Customer block: rows 8-17, contiguous, no section rows in the column.
        crr = raw.get("customer_row_range")
        if not crr or crr.get("first_row") != 8 or crr.get("last_row") != 17:
            failures.append(f"Raw Data: customer_row_range expected rows 8-17, got {crr!r}")
        if crr and not crr.get("contiguous"):
            failures.append("Raw Data: customer block expected contiguous")
        if raw.get("section_rows_in_customer_col"):
            failures.append(
                f"Raw Data: expected no section rows in customer column, got "
                f"{raw.get('section_rows_in_customer_col')!r}"
            )
        # Fixture is historical (ends Jun-25/26) → no forecast columns.
        if (raw.get("forecast") or {}).get("future_columns"):
            failures.append(
                f"Raw Data: expected no forecast columns, got "
                f"{(raw.get('forecast') or {}).get('future_columns')!r}"
            )
        if raw["sufficiency"]["verdict"] != "pass":
            failures.append(f"Raw Data: sufficiency expected 'pass', got "
                            f"{raw['sufficiency']['verdict']!r}")

    # --- Corkscrew assertions (two-pass behavior) --------------------------
    # Corkscrew is a finished-analysis sheet, not source data. With two-pass
    # survey it should be SKIPPED at Pass 1 (not deep-inspected) because Raw
    # Data dominates on score. The skipped list always records it; near_tie
    # should be False because the gap is wide.
    if "Corkscrew" in sheets_by_name:
        failures.append(
            "Corkscrew was deep-inspected — should have been skipped at Pass 1 "
            "in favor of Raw Data."
        )
    skipped_names = {s["name"] for s in report.get("skipped_sheets", [])}
    if "Corkscrew" not in skipped_names:
        failures.append("Corkscrew not present in skipped_sheets — Pass 1 missed it.")
    # --- Notes sheet: also expected to be skipped --------------------------
    if "Notes" in sheets_by_name:
        failures.append(
            "Notes sheet was deep-inspected — it has only one text cell and "
            "should appear in skipped_sheets."
        )
    if "Notes" not in skipped_names:
        failures.append("Notes not present in skipped_sheets.")
    if report.get("near_tie", False):
        failures.append(
            "near_tie expected False (Raw Data clearly dominates), "
            f"got True. Scores: {report.get('skipped_sheets')}"
        )

    # --- Overall ----------------------------------------------------------
    if report["overall_sufficiency"]["verdict"] != "pass":
        failures.append(
            f"Overall sufficiency expected 'pass', got "
            f"{report['overall_sufficiency']['verdict']!r}"
        )

    # =====================================================================
    # ADVERSARIAL BATTERY — exercises the customer-block detector and the
    # forecast/current-month logic against shape variations the single bundled
    # fixture does NOT cover. Each case is a compact spec; we build a sheet,
    # run inspect_sheet, and assert. The point is to prove the heuristics
    # generalise, not just fit the one workbook that prompted them.
    # =====================================================================
    def D(y, m):
        return dt.date(y, m, 1)

    FAR = dt.date(2099, 12, 1)  # default as_of → every test month is a complete actual

    def _build_battery_sheet(spec):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = spec["name"][:31]
        ws.cell(row=1, column=1, value="Customer")
        for j, d in enumerate(spec["dates"]):
            ws.cell(row=1, column=2 + j, value=d)
        for i, (label, vals) in enumerate(spec["rows"]):
            r = 2 + i
            if label is not None:
                ws.cell(row=r, column=1, value=label)
            for j, v in enumerate(vals or []):
                if v is not None:
                    ws.cell(row=r, column=2 + j, value=v)
        return ws

    M3 = [D(2025, 1), D(2025, 2), D(2025, 3)]
    M2 = [D(2025, 1), D(2025, 2)]

    battery = [
        # 1. Clean list, no summary, no blanks.
        {"name": "Clean", "dates": M3, "rows": [
            ("Acme", [1000, 1100, 1200]), ("Globex", [500, 600, 600]),
            ("Initech", [1500, 1400, 1450]), ("Umbrella", [300, 300, 350]),
        ], "expect": {"first": 2, "last": 5, "count": 4, "excluded": 0, "sections": set()}},

        # 2. Summary block ABOVE + blank gap, with a verbose unmatched line (the
        #    real-world shape that prompted this work).
        {"name": "SummaryAbove", "dates": M3, "rows": [
            ("Summary", [3000, 3100, 3200]),
            ("Pipeline to be signed / unidentified changes", [0, -250, 0]),
            ("Total", [3000, 2850, 3200]),
            (None, None), (None, None),
            ("Acme", [1000, 1100, 1200]), ("Globex", [500, 600, 600]),
            ("Initech", [1500, 1150, 1400]),
        ], "expect": {"first": 7, "last": 9, "count": 3, "excluded_min": 1,
                       "sections_superset": {"Summary", "Total"}, "cust_neg": 0}},

        # 3. Summary block BELOW + blank gap.
        {"name": "SummaryBelow", "dates": M2, "rows": [
            ("Acme", [1000, 1100]), ("Globex", [500, 600]), ("Initech", [1500, 1400]),
            (None, None),
            ("Total", [3000, 3100]), ("Memo: pipeline", [200, 200]),
        ], "expect": {"first": 2, "last": 4, "count": 3, "excluded_min": 1,
                       "sections_superset": {"Total"}}},

        # 4. Summary clusters at BOTH ends.
        {"name": "SummaryBothEnds", "dates": M2, "rows": [
            ("Total", [9999, 9999]), ("Memo top", [100, 100]),
            (None, None),
            ("Acme", [1000, 1100]), ("Globex", [500, 600]),
            ("Initech", [1500, 1400]), ("Umbrella", [300, 350]),
            (None, None),
            ("Memo bottom", [50, 50]),
        ], "expect": {"first": 5, "last": 8, "count": 4, "excluded": 2}},

        # 5. Summary directly above, NO blank, but RECOGNISED labels → sections.
        {"name": "NoBlankLabeled", "dates": M2, "rows": [
            ("Total MRR", [3000, 3100]), ("Summary", [3000, 3100]),
            ("Acme", [1000, 1100]), ("Globex", [500, 600]), ("Initech", [1500, 1400]),
        ], "expect": {"first": 4, "last": 6, "count": 3, "excluded": 0,
                       "sections_superset": {"Total MRR", "Summary"}}},

        # 6. Hidden TOTAL at top, NO blank, VERBOSE unmatched label → caught by
        #    the structural summation guard (signal E), not the vocabulary.
        {"name": "HiddenTotalTop", "dates": M2, "rows": [
            ("Consolidated book (all accounts)", [3000, 3100]),
            ("Acme", [1000, 1100]), ("Globex", [500, 600]), ("Initech", [1500, 1400]),
        ], "expect": {"first": 3, "last": 5, "count": 3,
                       "total_label": "Consolidated book (all accounts)"}},

        # 7. Hidden TOTAL at bottom, NO blank, verbose label.
        {"name": "HiddenTotalBottom", "dates": M2, "rows": [
            ("Acme", [1000, 1100]), ("Globex", [500, 600]), ("Initech", [1500, 1400]),
            ("Roll-up of the above", [3000, 3100]),
        ], "expect": {"first": 2, "last": 4, "count": 3,
                       "total_label": "Roll-up of the above"}},

        # 8. Section rows interspersed AMONG customers (no blanks). "New Relic"
        #    must NOT be read as a 'new' movement row; negatives tagged correctly.
        {"name": "Interspersed", "dates": M3, "rows": [
            ("Acme", [1000, -100, 1100]),       # customer-row negative
            ("Globex", [500, 600, 600]),
            ("New Relic", [900, 900, 900]),     # keyword-like name, real customer
            ("Churn", [0, -500, 0]),            # section-row negative
            ("Contraction", [0, -300, 0]),      # section-row negative
        ], "expect": {"count": 3, "sections_superset": {"Churn", "Contraction"},
                       "not_sections": {"New Relic"}, "cust_neg": 1, "sect_neg": 2}},

        # 9. RISK CASE: a real customer list split by a stray internal blank row.
        #    Must MERGE (not lop off half), reported non-contiguous, none excluded.
        {"name": "InternalBlank", "dates": M2, "rows": [
            ("Acme", [1000, 1100]), ("Bravo", [500, 600]), ("Cosmo", [1500, 1400]),
            (None, None),
            ("Dover", [800, 800]), ("Echo", [700, 700]), ("Foxtrot", [600, 600]),
        ], "expect": {"first": 2, "last": 8, "count": 6, "excluded": 0,
                       "contiguous": False}},

        # 10. Customer names that START with summary/movement keywords. None may
        #     be misread as a section (tests whole-cell label matching).
        {"name": "KeywordNames", "dates": M2, "rows": [
            ("Total Wine & More", [1000, 1100]), ("Sum Ventures", [500, 600]),
            ("New Relic", [300, 400]), ("Net Health", [200, 250]),
        ], "expect": {"count": 4, "sections": set()}},

        # 11. Single customer.
        {"name": "SingleCustomer", "dates": M3, "rows": [
            ("OnlyCo", [1000, 1100, 1200]),
        ], "expect": {"count": 1, "first": 2, "last": 2}},

        # 12. Current-month boundary: as_of mid-March → Mar is in-progress, last
        #     complete actual is Feb.
        {"name": "CurrentMonth", "dates": [D(2026, 1), D(2026, 2), D(2026, 3)],
         "as_of": dt.date(2026, 3, 10), "rows": [
            ("Acme", [1000, 1100, 1200]), ("Globex", [500, 600, 600]),
            ("Initech", [1500, 1400, 1450]),
        ], "expect": {"count": 3, "actuals_through": "Feb-26", "future_count": 1}},

        # 13. Forecast tail + current month (the user's scenario): as_of Jun-2026,
        #     data runs to Dec-2026 → last complete actual is May-26; Jun + Jul..Dec
        #     are not complete (7 columns).
        {"name": "ForecastTail",
         "dates": [D(2026, m) for m in range(1, 13)],
         "as_of": dt.date(2026, 6, 15), "rows": [
            ("Acme", [1000] * 12), ("Globex", [500] * 12), ("Initech", [1500] * 12),
        ], "expect": {"count": 3, "actuals_through": "May-26", "future_count": 7}},

        # 14. All months strictly in the past → no incomplete columns.
        {"name": "AllPast", "dates": [D(2023, 1), D(2023, 2), D(2023, 3)],
         "as_of": dt.date(2026, 6, 15), "rows": [
            ("Acme", [1000, 1100, 1200]), ("Globex", [500, 600, 600]),
            ("Initech", [1500, 1400, 1450]),
        ], "expect": {"count": 3, "actuals_through": "Mar-23", "future_count": 0}},

        # 15. Dataless label row (segment sub-header) between header and the first
        #     real customer — must NOT anchor the block.
        {"name": "DatalessLeadLabel", "dates": M3, "rows": [
            ("Enterprise", None),
            ("Acme", [1000, 1100, 1200]), ("Globex", [500, 600, 600]),
            ("Initech", [1500, 1400, 1450]),
        ], "expect": {"first": 3, "last": 5, "count": 3}},
    ]

    for spec in battery:
        nm = spec["name"]
        ws_b = _build_battery_sheet(spec)
        rep = inspect_sheet(ws_b, as_of=spec.get("as_of", FAR))
        exp = spec["expect"]
        crr = rep.get("customer_row_range") or {}
        secs = {x["label"] for x in rep.get("section_rows_in_customer_col", [])}
        excl = rep.get("excluded_row_clusters") or []
        fc = rep.get("forecast") or {}
        negs = rep.get("negative_values") or []

        if "count" in exp and rep.get("customer_count_hypothesis") != exp["count"]:
            failures.append(f"[{nm}] count expected {exp['count']}, got {rep.get('customer_count_hypothesis')}")
        if "first" in exp and crr.get("first_row") != exp["first"]:
            failures.append(f"[{nm}] first_row expected {exp['first']}, got {crr.get('first_row')}")
        if "last" in exp and crr.get("last_row") != exp["last"]:
            failures.append(f"[{nm}] last_row expected {exp['last']}, got {crr.get('last_row')}")
        if "contiguous" in exp and crr.get("contiguous") != exp["contiguous"]:
            failures.append(f"[{nm}] contiguous expected {exp['contiguous']}, got {crr.get('contiguous')}")
        if "excluded" in exp and len(excl) != exp["excluded"]:
            failures.append(f"[{nm}] excluded clusters expected {exp['excluded']}, got {len(excl)} ({excl})")
        if "excluded_min" in exp and len(excl) < exp["excluded_min"]:
            failures.append(f"[{nm}] excluded clusters expected >= {exp['excluded_min']}, got {len(excl)}")
        if "sections" in exp and secs != exp["sections"]:
            failures.append(f"[{nm}] sections expected {exp['sections']}, got {secs}")
        if "sections_superset" in exp and not exp["sections_superset"].issubset(secs):
            failures.append(f"[{nm}] sections missing {exp['sections_superset'] - secs} (got {secs})")
        if "not_sections" in exp and (exp["not_sections"] & secs):
            failures.append(f"[{nm}] these were wrongly classified as sections: {exp['not_sections'] & secs}")
        if "total_label" in exp:
            total_secs = {x["label"] for x in rep.get("section_rows_in_customer_col", []) if x.get("reason")}
            if exp["total_label"] not in total_secs:
                failures.append(f"[{nm}] expected total-row {exp['total_label']!r} flagged via summation, got {total_secs}")
        if "cust_neg" in exp:
            got = len([n for n in negs if n.get("in_customer_row")])
            if got != exp["cust_neg"]:
                failures.append(f"[{nm}] customer-row negatives expected {exp['cust_neg']}, got {got}")
        if "sect_neg" in exp:
            got = len([n for n in negs if not n.get("in_customer_row")])
            if got != exp["sect_neg"]:
                failures.append(f"[{nm}] section-row negatives expected {exp['sect_neg']}, got {got}")
        if "actuals_through" in exp and fc.get("actuals_through") != exp["actuals_through"]:
            failures.append(f"[{nm}] actuals_through expected {exp['actuals_through']}, got {fc.get('actuals_through')}")
        if "future_count" in exp and len(fc.get("future_columns", [])) != exp["future_count"]:
            failures.append(f"[{nm}] future/incomplete cols expected {exp['future_count']}, got {len(fc.get('future_columns', []))}")

    print(f"[battery] ran {len(battery)} adversarial cases")
    print("=" * 78)
    if failures:
        print(f"SELF-TEST: FAIL  ({len(failures)} assertion(s) failed)")
        for f in failures:
            print(f"  - {f}")
        print("=" * 78)
        return 1
    else:
        print("SELF-TEST: PASS")
        print("  Verified:")
        print("    - Sheet 'Raw Data': role=source data, shape=sectioned-wide, customer col B")
        print("    - Date columns C-T (18 months, Jan-25 to Jun-26)")
        print("    - Scale verdict starts with 'MRR'")
        if raw:
            print(f"    - {len(raw['derived_blocks_detected'])} derived blocks detected (New/Upsell/Downsell/Churn/Check)")
        print(f"    - Sheet 'Corkscrew': skipped at Pass 1 (score gap vs. Raw Data)")
        print(f"    - Sheet 'Notes': skipped (no date-like header row)")
        print(f"    - near_tie = {report.get('near_tie')}")
        print(f"    - Overall sufficiency = {report['overall_sufficiency']['verdict']}")
        print("=" * 78)
        return 0


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

def build_deliver_config(report: dict) -> dict | None:
    """Pick the 'source data' sheet and emit the exact params deliver.py needs,
    so its findings flow straight into the delivery step (no hand-copied flags)."""
    src = next(
        (s for s in report.get("sheets", [])
         if s.get("hypothesis", {}).get("role") == "source data"),
        None,
    )
    if not src:
        return None
    cc = src.get("candidate_customer_column") or {}
    dc = src.get("candidate_date_columns") or {}
    crr = src.get("customer_row_range") or {}
    fc = src.get("forecast") or {}
    return {
        "source_sheet": src.get("name"),
        "source_customer_col": cc.get("col_letter"),
        "source_first_data_row": crr.get("first_row"),
        "source_last_data_row": crr.get("last_row"),
        "source_first_date_col": dc.get("first_col_letter"),
        "source_header_row": dc.get("header_row"),
        "actuals_through": fc.get("actuals_through"),
    }


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        description="Retention-analysis Phase 1-4 inspector. Produces hypotheses; "
                    "does NOT extract data into a working dataset."
    )
    p.add_argument("path", nargs="?", default=None,
                   help="Path to .xlsx or .csv input file. Optional when "
                        "--self-test is supplied (defaults to the bundled fixture).")
    p.add_argument("--json", action="store_true",
                   help="Emit machine-readable JSON instead of human text")
    p.add_argument("--self-test", action="store_true",
                   help="Run built-in self-test against the bundled sample fixture "
                        "(or a supplied path, if given)")
    p.add_argument("--emit-config", metavar="PATH",
                   help="Write the source-sheet params deliver.py needs to a JSON "
                        "file (feed it back via deliver.py --config PATH)")
    args = p.parse_args(argv)

    if args.self_test:
        return run_self_test(args.path)

    if args.path is None:
        print("ERROR: path is required (or use --self-test)", file=sys.stderr)
        return 2

    if not os.path.exists(args.path):
        print(f"ERROR: file not found: {args.path}", file=sys.stderr)
        return 2

    _t0 = time.perf_counter()
    report = inspect_file(args.path)
    if args.json:
        # JSON-safe: convert any datetime customer values, etc.
        def default(o):
            if isinstance(o, (dt.date, dt.datetime)):
                return o.isoformat()
            raise TypeError(f"not JSON-serializable: {type(o).__name__}")
        print(json.dumps(report, indent=2, default=default))
    else:
        print(render_text(report))
    if args.emit_config:
        cfg = build_deliver_config(report)
        if cfg is None:
            print("WARN: no 'source data' sheet found; no config written.",
                  file=sys.stderr)
        else:
            with open(args.emit_config, "w", encoding="utf-8") as fh:
                json.dump(cfg, fh, indent=2)
            print(f"[survey.py] wrote deliver config -> {args.emit_config}",
                  file=sys.stderr)
    print(f"[survey.py] inspected in {time.perf_counter() - _t0:.2f}s", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
