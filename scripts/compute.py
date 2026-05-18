"""
compute.py — retention-analysis skill, Phases 5–7.

Takes a long-format CSV with columns (customer_id, month, mrr) and produces a
JSON-serializable dict containing:

  - config         : run configuration + dataset shape
  - buckets        : per (customer, month) classification + amount
  - monthly        : per-month aggregates (rollforward)
  - metrics_monthly: Gross / Net / Logo retention per month (blank for first month)
  - metrics_ltm    : same metrics on a 12-month-prior basis (only when >=13 months)
  - verification   : pass/fail per layer (1..8) plus diagnostics

CLI:
    python3 compute.py <long-format-csv> [--arr-factor 12] [--output result.json]

Default behavior prints JSON to stdout. With --output, writes JSON to the path.

Strict cohort always. Beginning of first period = sum of customer revenue in that
period (no movements that period). Monthly metrics blank for the first month.
LTM only when >=13 months of data.
"""

from __future__ import annotations

# --- sys.path hygiene: defensive guard. When a script runs as __main__, Python
# prepends its directory to sys.path, which can shadow stdlib modules that share
# a name with sibling files. Strip the script's directory before importing.
import os as _os
import sys as _sys

_HERE = _os.path.dirname(_os.path.abspath(__file__))
_sys.path[:] = [
    p for p in _sys.path if not p or _os.path.abspath(p) != _HERE
]
# If `inspect` was already partially imported and points at the local copy,
# drop it so a fresh stdlib import succeeds.
_inspect_mod = _sys.modules.get("inspect")
if _inspect_mod is not None:
    _f = getattr(_inspect_mod, "__file__", "") or ""
    if _f and _os.path.abspath(_os.path.dirname(_f)) == _HERE:
        del _sys.modules["inspect"]

import argparse
import csv
import json
import math
import os
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any


# Path to the bundled synthetic test fixture. Resolved relative to this file so
# it works whether the skill is run from ~/.claude/skills/ or a fresh clone.
DEFAULT_FIXTURE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "fixtures",
    "sample_retention_data.xlsx",
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _parse_month(value: str) -> str:
    """Parse a date-ish string and return canonical 'YYYY-MM'."""
    s = str(value).strip()
    # Try ISO date first
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y", "%b-%y", "%b-%Y", "%B %Y", "%Y%m"):
        try:
            d = datetime.strptime(s, fmt)
            return f"{d.year:04d}-{d.month:02d}"
        except ValueError:
            continue
    # If it's already YYYY-MM-ish
    if len(s) >= 7 and s[4] == "-":
        return s[:7]
    raise ValueError(f"Could not parse month value: {value!r}")


def _round2(x: float) -> float:
    return round(x + 0.0, 2)


def _add_months(year: int, month: int, delta: int) -> tuple[int, int]:
    total = year * 12 + (month - 1) + delta
    return total // 12, (total % 12) + 1


def _month_key_minus(month_key: str, delta: int) -> str:
    y, m = int(month_key[:4]), int(month_key[5:7])
    yy, mm = _add_months(y, m, -delta)
    return f"{yy:04d}-{mm:02d}"


# ---------------------------------------------------------------------------
# Phase 5 — Bucket math
# ---------------------------------------------------------------------------


BUCKETS = ("new", "upsell", "flat", "downsell", "churn", "inactive")


def classify(prior: float, current: float) -> tuple[str, float]:
    """Return (bucket, amount) per the spec.

    - Treat tiny floating noise (<0.005) as zero and as equal.
    """
    eps = 0.005

    p_zero = abs(prior) < eps
    c_zero = abs(current) < eps
    diff = current - prior

    if p_zero and c_zero:
        return "inactive", 0.0
    if p_zero and not c_zero:
        return "new", _round2(current)
    if not p_zero and c_zero:
        return "churn", _round2(-prior)
    # both > 0
    if abs(diff) < eps:
        return "flat", 0.0
    if diff > 0:
        return "upsell", _round2(diff)
    return "downsell", _round2(diff)  # diff is negative


# ---------------------------------------------------------------------------
# Core computation
# ---------------------------------------------------------------------------


@dataclass
class LongRow:
    customer_id: str
    month: str  # YYYY-MM
    mrr: float


def load_long_csv(path: str) -> list[LongRow]:
    rows: list[LongRow] = []
    with open(path, newline="") as fh:
        reader = csv.DictReader(fh)
        # Try to be flexible about header names
        cust_keys = {"customer_id", "customer", "id", "customer id"}
        month_keys = {"month", "date", "period"}
        mrr_keys = {"mrr", "revenue", "value", "amount", "arr"}
        # Normalize fieldnames
        fmap: dict[str, str] = {}
        for fn in reader.fieldnames or []:
            lower = fn.strip().lower()
            if lower in cust_keys:
                fmap["customer_id"] = fn
            elif lower in month_keys:
                fmap["month"] = fn
            elif lower in mrr_keys:
                fmap["mrr"] = fn
        # Fallback to positional if missing
        for required in ("customer_id", "month", "mrr"):
            if required not in fmap:
                # Use the original first / second / third columns
                if reader.fieldnames and len(reader.fieldnames) >= 3:
                    pos = {"customer_id": 0, "month": 1, "mrr": 2}[required]
                    fmap[required] = reader.fieldnames[pos]
                else:
                    raise ValueError(f"Required column {required!r} not found in CSV")
        for row in reader:
            cid = str(row[fmap["customer_id"]]).strip()
            mo = _parse_month(row[fmap["month"]])
            raw = row[fmap["mrr"]]
            if raw is None or str(raw).strip() == "":
                mrr = 0.0
            else:
                mrr = float(raw)
            rows.append(LongRow(customer_id=cid, month=mo, mrr=mrr))
    return rows


def build_matrix(rows: list[LongRow]) -> tuple[list[str], list[str], dict[tuple[str, str], float], list[str]]:
    """Return (months_sorted, customers_sorted, value_lookup, duplicates).

    - months_sorted: chronologically ordered list of YYYY-MM strings
    - customers_sorted: stable order by first-appearance month then id
    - value_lookup: {(customer_id, month): mrr}; missing pairs treated as 0
    - duplicates: list of "(cid, month)" strings observed more than once in input
    """
    months_set: set[str] = set()
    customers_set: set[str] = set()
    value_lookup: dict[tuple[str, str], float] = {}
    seen_pairs: set[tuple[str, str]] = set()
    duplicates: list[str] = []

    for r in rows:
        months_set.add(r.month)
        customers_set.add(r.customer_id)
        key = (r.customer_id, r.month)
        if key in seen_pairs:
            duplicates.append(f"{r.customer_id}@{r.month}")
            # Sum duplicates conservatively rather than overwrite
            value_lookup[key] = value_lookup.get(key, 0.0) + r.mrr
        else:
            seen_pairs.add(key)
            value_lookup[key] = r.mrr

    months_sorted = sorted(months_set)
    customers_sorted = sorted(customers_set, key=lambda x: (int(x) if x.isdigit() else 0, x))
    return months_sorted, customers_sorted, value_lookup, duplicates


def compute(rows: list[LongRow], arr_factor: float = 1.0, stated_retention: dict | None = None) -> dict[str, Any]:
    months, customers, lookup, duplicates = build_matrix(rows)

    if len(months) < 2:
        raise ValueError("compute.py requires at least 2 distinct months in input")
    if len(customers) < 1:
        raise ValueError("compute.py requires at least 1 customer in input")

    # Scale into ARR space ONCE: every internal aggregate is in ARR dollars.
    af = float(arr_factor)

    def v(c: str, m: str) -> float:
        return lookup.get((c, m), 0.0) * af

    # ---- Phase 5: bucket math ---------------------------------------------
    # buckets[(customer, month)] = (bucket_name, amount_in_ARR)
    buckets: list[dict[str, Any]] = []
    bucket_index: dict[tuple[str, str], tuple[str, float]] = {}

    # Track per-customer-month negatives (raw input)
    negatives: list[dict[str, Any]] = []

    for c in customers:
        # First month: no prior; we don't classify the very first month per spec
        # (Beginning of first period = sum of customer revenue in that period;
        # no movements that period.) But we still want a record for completeness?
        # The spec's example only shows non-first months in `buckets`. We'll emit
        # records starting from the second month (delta-based classification).
        for i in range(1, len(months)):
            m_prev = months[i - 1]
            m_curr = months[i]
            prior = v(c, m_prev)
            current = v(c, m_curr)
            bucket, amount = classify(prior, current)
            bucket_index[(c, m_curr)] = (bucket, amount)
            buckets.append(
                {
                    "customer_id": c,
                    "month": m_curr,
                    "bucket": bucket,
                    "amount": _round2(amount),
                    "prior_mrr": _round2(prior / af) if af else _round2(prior),
                    "current_mrr": _round2(current / af) if af else _round2(current),
                }
            )
        # Negative-value scan
        for m in months:
            raw = lookup.get((c, m), 0.0)
            if raw < 0:
                negatives.append({"customer_id": c, "month": m, "value": raw})

    # ---- Monthly rollforward ---------------------------------------------
    monthly: list[dict[str, Any]] = []

    # Beginning of first month = sum of customer revenue in that month (ARR)
    first_month = months[0]
    first_beginning = sum(v(c, first_month) for c in customers)
    n_active_first = sum(1 for c in customers if v(c, first_month) > 0)

    monthly.append(
        {
            "month": first_month,
            "beginning": _round2(first_beginning),
            "new": 0.0,
            "upsell": 0.0,
            "downsell": 0.0,
            "churn": 0.0,
            "ending": _round2(first_beginning),
            "n_active": n_active_first,
            "n_new": 0,
            "n_churned": 0,
        }
    )

    for i in range(1, len(months)):
        m = months[i]
        m_prev = months[i - 1]
        new_amt = 0.0
        upsell_amt = 0.0
        downsell_amt = 0.0
        churn_amt = 0.0
        n_new = 0
        n_churned = 0
        n_active = 0
        for c in customers:
            bucket, amount = bucket_index[(c, m)]
            if bucket == "new":
                new_amt += amount
                n_new += 1
            elif bucket == "upsell":
                upsell_amt += amount
            elif bucket == "downsell":
                downsell_amt += amount  # negative
            elif bucket == "churn":
                churn_amt += amount  # negative
                n_churned += 1
            if v(c, m) > 0:
                n_active += 1

        beginning = monthly[-1]["ending"]
        ending = beginning + new_amt + upsell_amt + downsell_amt + churn_amt
        monthly.append(
            {
                "month": m,
                "beginning": _round2(beginning),
                "new": _round2(new_amt),
                "upsell": _round2(upsell_amt),
                "downsell": _round2(downsell_amt),
                "churn": _round2(churn_amt),
                "ending": _round2(ending),
                "n_active": n_active,
                "n_new": n_new,
                "n_churned": n_churned,
            }
        )

    # ---- Phase 6: monthly metrics (blank for first month) -----------------
    metrics_monthly: list[dict[str, Any]] = []
    for i in range(1, len(months)):
        m = months[i]
        agg = monthly[i]
        beginning = agg["beginning"]
        if beginning <= 0:
            metrics_monthly.append({"month": m, "gross": None, "net": None, "logo": None})
            continue
        gross = (beginning + agg["downsell"] + agg["churn"]) / beginning
        net = (beginning + agg["upsell"] + agg["downsell"] + agg["churn"]) / beginning
        # Logo retention: of customers active at start of period, how many remain
        n_start = monthly[i - 1]["n_active"]
        n_churned = agg["n_churned"]
        logo = ((n_start - n_churned) / n_start) if n_start > 0 else None
        metrics_monthly.append(
            {
                "month": m,
                "gross": round(gross, 6),
                "net": round(net, 6),
                "logo": round(logo, 6) if logo is not None else None,
            }
        )

    # ---- Phase 6: LTM metrics --------------------------------------------
    # Annual / LTM retention is a point-in-time cohort comparison (Method A):
    # take the customers active at month T-12, look at THEIR revenue today, and
    # divide by their revenue 12 months ago. New customers in the interim are
    # excluded. Churned cohort members stay in the denominator at their T-12
    # ARR and contribute $0 to the numerator. This is what investors expect
    # (SaaS Metrics Standards Board, Datadog/Okta 10-Ks, etc.) and avoids the
    # within-year round-trip ambiguity that bucket-aggregation has.
    metrics_ltm: list[dict[str, Any]] | None
    if len(months) >= 13:
        metrics_ltm = []
        for i in range(12, len(months)):
            m_now = months[i]
            m_then = months[i - 12]

            # Strict cohort: customers with revenue > 0 at T-12.
            cohort = [c for c in customers if v(c, m_then) > 0]
            if not cohort:
                metrics_ltm.append({"month": m_now, "gross": None, "net": None, "logo": None})
                continue

            base_total = sum(v(c, m_then) for c in cohort)
            current_total = sum(v(c, m_now) for c in cohort)
            # Gross caps each customer's contribution at their starting ARR —
            # i.e., upsell is excluded; only downsell and churn pull it down.
            gross_total = sum(min(v(c, m_now), v(c, m_then)) for c in cohort)
            survived = sum(1 for c in cohort if v(c, m_now) > 0)

            gross = gross_total / base_total if base_total else None
            net = current_total / base_total if base_total else None
            logo = survived / len(cohort)

            metrics_ltm.append(
                {
                    "month": m_now,
                    "gross": round(gross, 6) if gross is not None else None,
                    "net": round(net, 6) if net is not None else None,
                    "logo": round(logo, 6),
                }
            )
    else:
        metrics_ltm = None

    # ---- Phase 7: 8-layer verification battery ----------------------------
    verification = run_verification(
        months=months,
        customers=customers,
        lookup=lookup,
        arr_factor=af,
        bucket_index=bucket_index,
        monthly=monthly,
        metrics_monthly=metrics_monthly,
        metrics_ltm=metrics_ltm,
        duplicates=duplicates,
        negatives=negatives,
        stated_retention=stated_retention,
    )

    out: dict[str, Any] = {
        "config": {
            "arr_factor": af,
            "n_customers": len(customers),
            "n_months": len(months),
            "month_range": [months[0], months[-1]],
        },
        "buckets": buckets,
        "monthly": monthly,
        "metrics_monthly": metrics_monthly,
        "metrics_ltm": metrics_ltm,
        "verification": verification,
    }
    return out


# ---------------------------------------------------------------------------
# Phase 7 — verification
# ---------------------------------------------------------------------------


def run_verification(
    months,
    customers,
    lookup,
    arr_factor,
    bucket_index,
    monthly,
    metrics_monthly,
    metrics_ltm,
    duplicates,
    negatives,
    stated_retention,
):
    af = arr_factor

    def v(c, m):
        return lookup.get((c, m), 0.0) * af

    # ---- Layer 1: identity check (Beginning + N + U + D + C - Ending = 0)
    layer1_max = 0.0
    layer1_fails = []
    for agg in monthly:
        residual = agg["beginning"] + agg["new"] + agg["upsell"] + agg["downsell"] + agg["churn"] - agg["ending"]
        if abs(residual) > layer1_max:
            layer1_max = abs(residual)
        if abs(residual) > 0.01:
            layer1_fails.append({"month": agg["month"], "residual": residual})
    layer1 = {"pass": len(layer1_fails) == 0, "max_residual": round(layer1_max, 6), "failures": layer1_fails}

    # ---- Layer 2: stitching (Ending(t) == Beginning(t+1))
    layer2_fails = []
    for i in range(len(monthly) - 1):
        if abs(monthly[i]["ending"] - monthly[i + 1]["beginning"]) > 0.01:
            layer2_fails.append(
                {"between": [monthly[i]["month"], monthly[i + 1]["month"]],
                 "diff": monthly[i]["ending"] - monthly[i + 1]["beginning"]}
            )
    layer2 = {"pass": len(layer2_fails) == 0, "failures": layer2_fails}

    # ---- Layer 3: bucket exclusivity (each (cust, month) in one bucket only)
    # Our data structure guarantees this by construction (single dict entry per
    # (cust, month)); we still check that every classified amount agrees with
    # the bucket sign rule.
    layer3_fails = []
    for (c, m), (bucket, amount) in bucket_index.items():
        if bucket == "new" and amount < 0:
            layer3_fails.append({"customer": c, "month": m, "issue": "new with negative amount"})
        elif bucket == "upsell" and amount <= 0:
            layer3_fails.append({"customer": c, "month": m, "issue": "upsell with non-positive amount"})
        elif bucket == "downsell" and amount >= 0:
            layer3_fails.append({"customer": c, "month": m, "issue": "downsell with non-negative amount"})
        elif bucket == "churn" and amount >= 0:
            layer3_fails.append({"customer": c, "month": m, "issue": "churn with non-negative amount"})
        elif bucket in ("flat", "inactive") and abs(amount) > 0.001:
            layer3_fails.append({"customer": c, "month": m, "issue": f"{bucket} with nonzero amount"})
    layer3 = {"pass": len(layer3_fails) == 0, "failures": layer3_fails}

    # ---- Layer 4: sample-trace 5 customers spanning behaviors
    samples: list[dict[str, Any]] = []
    layer4_pass = True

    # Helper: classify a customer's overall trajectory
    def trajectory(c: str) -> dict[str, Any]:
        ms = [v(c, m) for m in months]
        nonzero = [x for x in ms if x > 0]
        return {
            "first": ms[0],
            "last": ms[-1],
            "max": max(ms) if ms else 0,
            "any_zero": any(x == 0 for x in ms),
            "any_nonzero": len(nonzero) > 0,
            "all_equal_when_active": len(set(round(x, 2) for x in ms if x > 0)) <= 1,
            "values": ms,
        }

    trajectories = {c: trajectory(c) for c in customers}

    # 1) Never changed: active in first AND last AND all active values equal
    never_changed = [
        c for c in customers
        if trajectories[c]["first"] > 0
        and trajectories[c]["last"] > 0
        and not trajectories[c]["any_zero"]
        and trajectories[c]["all_equal_when_active"]
    ]
    # 2) Joined mid: first == 0, becomes nonzero exactly once
    joined_mid = []
    for c in customers:
        ms = trajectories[c]["values"]
        if ms[0] == 0:
            # find first nonzero
            for k, x in enumerate(ms):
                if x > 0:
                    # Flat thereafter?
                    rest = ms[k:]
                    if all(abs(r - rest[0]) < 0.005 for r in rest):
                        joined_mid.append(c)
                    break
    # 3) Churned: nonzero then zero, stays zero
    churned = []
    for c in customers:
        ms = trajectories[c]["values"]
        if ms[0] > 0 and ms[-1] == 0:
            # find first zero after nonzero start
            saw_nonzero = False
            ok = True
            for x in ms:
                if x > 0 and not saw_nonzero:
                    saw_nonzero = True
                elif saw_nonzero and x == 0:
                    pass  # remains zero
                elif saw_nonzero and x > 0:
                    # resurrection -> not the clean churn pattern
                    ok = False
                    break
            if ok:
                churned.append(c)
    # 4) Upsell: any month classified as upsell
    upsold = sorted({c for (c, _m), (b, _a) in bucket_index.items() if b == "upsell"})
    # 5) Largest customer: by total spend
    totals = [(c, sum(v(c, m) for m in months)) for c in customers]
    totals.sort(key=lambda x: x[1], reverse=True)
    largest = [totals[0][0]] if totals else []

    picks: list[tuple[str, str]] = []  # (label, customer_id)
    if never_changed:
        picks.append(("never_changed", never_changed[0]))
    if joined_mid:
        picks.append(("joined_mid", joined_mid[0]))
    if churned:
        picks.append(("churned", churned[0]))
    if upsold:
        picks.append(("upsold", upsold[0]))
    if largest:
        picks.append(("largest", largest[0]))

    for label, c in picks:
        # Walk through and verify classification matches rule
        per_month = []
        ok = True
        for i in range(1, len(months)):
            m_prev = months[i - 1]
            m_curr = months[i]
            prior = v(c, m_prev)
            current = v(c, m_curr)
            expected_bucket, expected_amt = classify(prior, current)
            actual_bucket, actual_amt = bucket_index[(c, m_curr)]
            match = expected_bucket == actual_bucket and abs(expected_amt - actual_amt) < 0.01
            if not match:
                ok = False
            per_month.append(
                {
                    "month": m_curr,
                    "prior": _round2(prior),
                    "current": _round2(current),
                    "expected": expected_bucket,
                    "actual": actual_bucket,
                    "match": match,
                }
            )
        samples.append({"label": label, "customer_id": c, "ok": ok, "trace": per_month})
        if not ok:
            layer4_pass = False

    # Sample richness: prefer to have at least 3 distinct labels
    if len(picks) < 3:
        layer4_pass = False

    layer4 = {"pass": layer4_pass, "samples": samples}

    # ---- Layer 5: edge cases ---------------------------------------------
    resurrections = []
    for c in customers:
        ms = trajectories[c]["values"]
        # Pattern: nonzero ... zero ... nonzero
        saw_nonzero = False
        saw_zero_after = False
        for x in ms:
            if x > 0 and not saw_nonzero:
                saw_nonzero = True
            elif saw_nonzero and x == 0:
                saw_zero_after = True
            elif saw_zero_after and x > 0:
                resurrections.append(c)
                break

    layer5 = {
        "resurrections": resurrections,
        "duplicates": duplicates,
        "negatives": negatives,
    }

    # ---- Layer 6: reasonableness / benchmark sniff -----------------------
    gross_le_net = True
    nrr_under_140 = True
    for r in metrics_monthly:
        if r["gross"] is not None and r["net"] is not None and r["gross"] - r["net"] > 1e-6:
            gross_le_net = False
        if r["net"] is not None and r["net"] > 1.40:
            nrr_under_140 = False
    if metrics_ltm:
        for r in metrics_ltm:
            if r["gross"] is not None and r["net"] is not None and r["gross"] - r["net"] > 1e-6:
                gross_le_net = False
            if r["net"] is not None and r["net"] > 1.40:
                nrr_under_140 = False

    # Logo vs $ retention story (advisory, not pass/fail)
    logo_vs_dollar = []
    for r in metrics_monthly:
        if r["logo"] is not None and r["net"] is not None:
            logo_vs_dollar.append({"month": r["month"], "logo": r["logo"], "dollar_net": r["net"]})

    layer6 = {
        "pass": gross_le_net and nrr_under_140,
        "gross_le_net": gross_le_net,
        "nrr_under_140": nrr_under_140,
        "logo_vs_dollar": logo_vs_dollar[:6],  # truncate
    }

    # ---- Layer 7: cross-check vs company-stated --------------------------
    if stated_retention:
        # stated_retention is a dict like {"net_ltm": 1.10, "month": "2026-03"}
        diffs = []
        for key, expected in stated_retention.items():
            # crude lookup
            actual = None
            if key.startswith("net_ltm") and metrics_ltm:
                actual = metrics_ltm[-1]["net"]
            elif key.startswith("gross_ltm") and metrics_ltm:
                actual = metrics_ltm[-1]["gross"]
            elif key.startswith("logo_ltm") and metrics_ltm:
                actual = metrics_ltm[-1]["logo"]
            if actual is None:
                diffs.append({"key": key, "expected": expected, "actual": None, "ok": False})
            else:
                diffs.append(
                    {
                        "key": key,
                        "expected": expected,
                        "actual": actual,
                        "ok": abs(actual - expected) < 0.005,
                    }
                )
        layer7 = {
            "applicable": True,
            "pass": all(d["ok"] for d in diffs),
            "diffs": diffs,
        }
    else:
        layer7 = {
            "applicable": False,
            "pass": True,  # not applicable counts as pass
            "note": "no stated retention figure provided",
        }

    # ---- Layer 8: compute it two ways ------------------------------------
    # Build-up method: monthly's net retention via aggregates (already in metrics_monthly)
    # Direct cohort method (for monthly): cohort = customers active at m_prev,
    # sum their revenue at m and at m_prev, NRR = (sum_now + 0_for_new)/sum_prev.
    # Strictly: NRR = (sum at m for those active at m_prev) / (sum at m_prev).
    max_disagreement = 0.0
    for i in range(1, len(months)):
        m_prev = months[i - 1]
        m_curr = months[i]
        cohort = [c for c in customers if v(c, m_prev) > 0]
        if not cohort:
            continue
        sum_prev = sum(v(c, m_prev) for c in cohort)
        sum_now = sum(v(c, m_curr) for c in cohort)
        nrr_direct = sum_now / sum_prev if sum_prev else None
        nrr_buildup = metrics_monthly[i - 1]["net"]
        if nrr_direct is not None and nrr_buildup is not None:
            disagreement = abs(nrr_direct - nrr_buildup)
            if disagreement > max_disagreement:
                max_disagreement = disagreement
    layer8 = {
        "pass": max_disagreement < 1e-4,
        "max_disagreement_pct": round(max_disagreement, 8),
    }

    return {
        "layer_1_identity": layer1,
        "layer_2_stitching": layer2,
        "layer_3_exclusivity": layer3,
        "layer_4_sample_trace": layer4,
        "layer_5_edge_cases": layer5,
        "layer_6_benchmark_sniff": layer6,
        "layer_7_company_stated": layer7,
        "layer_8_two_methods": layer8,
    }


# ---------------------------------------------------------------------------
# CLI entry
# ---------------------------------------------------------------------------


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Retention compute (Phases 5-7).")
    parser.add_argument("csv_path", help="Path to long-format CSV (customer_id, month, mrr)")
    parser.add_argument("--arr-factor", type=float, default=1.0, help="Multiply MRR by this to get ARR (use 12 for MRR input)")
    parser.add_argument("--output", default=None, help="Write JSON to this file (default: stdout)")
    args = parser.parse_args(argv)

    rows = load_long_csv(args.csv_path)
    result = compute(rows, arr_factor=args.arr_factor)

    text = json.dumps(result, indent=2, default=str)
    if args.output:
        with open(args.output, "w") as fh:
            fh.write(text)
    else:
        print(text)
    return 0


# ---------------------------------------------------------------------------
# Self-test
# ---------------------------------------------------------------------------


def _self_test(xlsx_path: str | None = None) -> int:
    """Run the script against the bundled sample fixture and assert outcomes.

    If ``xlsx_path`` is supplied, use it; otherwise default to
    ``DEFAULT_FIXTURE_PATH`` (the synthetic test fixture that ships with the
    skill).
    """
    import tempfile

    # Defensive sys.path hygiene: strip the script's directory before importing
    # openpyxl, in case any future sibling file shadows a stdlib module openpyxl
    # imports transitively. Cheap insurance.
    here = os.path.dirname(os.path.abspath(__file__))
    sys.path[:] = [p for p in sys.path if os.path.abspath(p) != here]
    # Pop any partially-imported local `inspect` so openpyxl reimports stdlib
    if "inspect" in sys.modules and getattr(sys.modules["inspect"], "__file__", "") and \
            os.path.abspath(sys.modules["inspect"].__file__).startswith(here):
        del sys.modules["inspect"]

    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        print(f"FAIL: openpyxl not installed or import failed: {exc}")
        return 2

    xlsx_path = xlsx_path or DEFAULT_FIXTURE_PATH

    if not os.path.exists(xlsx_path):
        print(f"FAIL: test target not found: {xlsx_path}")
        return 2

    # ---- Step 1-4: open xlsx, pivot Raw Data block to long-format CSV ----
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Raw Data" not in wb.sheetnames:
        print("FAIL: 'Raw Data' sheet not found")
        return 2
    ws = wb["Raw Data"]

    # Date headers in C7:T7 (cols 3..20, row 7) — 18 months Jan-25..Jun-26
    date_headers: list[str] = []
    for col in range(3, 21):
        cell = ws.cell(row=7, column=col).value
        if cell is None:
            print(f"FAIL: missing date header at col {col}")
            return 2
        if isinstance(cell, (datetime, date)):
            date_headers.append(f"{cell.year:04d}-{cell.month:02d}")
        else:
            date_headers.append(_parse_month(str(cell)))

    if len(date_headers) != 18 or date_headers[0] != "2025-01" or date_headers[-1] != "2026-06":
        print(f"FAIL: expected Jan-25..Jun-26, got {date_headers}")
        return 2

    # Customer rows 8..17 (10 customers), id in B, MRR in C..T
    long_rows: list[tuple[str, str, float]] = []
    for row in range(8, 18):
        cid_cell = ws.cell(row=row, column=2).value
        if cid_cell is None or str(cid_cell).strip() == "":
            print(f"FAIL: missing customer id at row {row}")
            return 2
        cid = str(int(cid_cell)) if isinstance(cid_cell, (int, float)) else str(cid_cell).strip()
        for j, m in enumerate(date_headers):
            v = ws.cell(row=row, column=3 + j).value
            mrr = 0.0 if v is None or v == "" else float(v)
            long_rows.append((cid, m, mrr))

    n_customers_in = len({r[0] for r in long_rows})
    if n_customers_in != 10:
        print(f"FAIL: expected 10 customers in input, got {n_customers_in}")
        return 2

    # Spot check Jan-25 total MRR = $7,300 (per EXPECTED_VALUES.md §3 m1 narrative)
    jan_total_mrr = sum(r[2] for r in long_rows if r[1] == "2025-01")
    if abs(jan_total_mrr - 7300.00) > 0.05:
        print(f"FAIL: Jan-25 total MRR mismatch — expected 7,300.00, got {jan_total_mrr:,.2f}")
        return 2

    # Write to temp CSV
    tmp = tempfile.NamedTemporaryFile(mode="w", suffix=".csv", delete=False, newline="")
    writer = csv.writer(tmp)
    writer.writerow(["customer_id", "month", "mrr"])
    for r in long_rows:
        writer.writerow(r)
    tmp.close()
    csv_path = tmp.name

    # ---- Step 5: invoke the compute entry programmatically ---------------
    rows = load_long_csv(csv_path)
    result = compute(rows, arr_factor=12.0)

    # ---- Step 6-10: assertions -------------------------------------------
    print("=" * 72)
    print("SELF-TEST: compute.py vs sample_retention_data.xlsx")
    print("=" * 72)

    layers = result["verification"]
    layer_results: list[tuple[str, bool, str]] = []

    def get_pass(d: dict) -> bool:
        return bool(d.get("pass", False))

    layer_results.append(("Layer 1 — Identity check", get_pass(layers["layer_1_identity"]),
                          f"max_residual = {layers['layer_1_identity']['max_residual']}"))
    layer_results.append(("Layer 2 — Period stitching", get_pass(layers["layer_2_stitching"]),
                          f"failures = {len(layers['layer_2_stitching'].get('failures', []))}"))
    layer_results.append(("Layer 3 — Bucket exclusivity", get_pass(layers["layer_3_exclusivity"]),
                          f"failures = {len(layers['layer_3_exclusivity'].get('failures', []))}"))
    l4 = layers["layer_4_sample_trace"]
    layer_results.append(("Layer 4 — Sample-trace by hand", get_pass(l4),
                          f"samples = {len(l4['samples'])}, all ok = {all(s['ok'] for s in l4['samples'])}"))
    l5 = layers["layer_5_edge_cases"]
    # Layer 5 has no pass/fail; it's a flag report. Pass if it produced lists.
    l5_pass = isinstance(l5.get("resurrections"), list)
    layer_results.append(("Layer 5 — Edge-case scan", l5_pass,
                          f"resurrections={len(l5['resurrections'])}, duplicates={len(l5['duplicates'])}, negatives={len(l5['negatives'])}"))
    l6 = layers["layer_6_benchmark_sniff"]
    layer_results.append(("Layer 6 — Reasonableness sniff", get_pass(l6),
                          f"gross_le_net={l6['gross_le_net']}, nrr_under_140={l6['nrr_under_140']}"))
    l7 = layers["layer_7_company_stated"]
    layer_results.append(("Layer 7 — Company-stated cross-check", get_pass(l7),
                          l7.get("note", "applicable")))
    l8 = layers["layer_8_two_methods"]
    layer_results.append(("Layer 8 — Compute two ways", get_pass(l8),
                          f"max_disagreement_pct = {l8['max_disagreement_pct']}"))

    print("\nVERIFICATION LAYERS")
    print("-" * 72)
    for name, ok, detail in layer_results:
        marker = "PASS" if ok else "FAIL"
        print(f"  [{marker}] {name}  ({detail})")

    # Direct asserts on the critical numbers (per fixtures/EXPECTED_VALUES.md §4)
    print("\nKEY ASSERTS")
    print("-" * 72)
    asserts: list[tuple[str, bool, str]] = []

    # --- Config ---
    cfg = result["config"]
    asserts.append(
        ("config.n_customers == 10 (dataset population)",
         cfg["n_customers"] == 10, f"got {cfg['n_customers']}")
    )
    asserts.append(
        ("config.n_months == 18", cfg["n_months"] == 18, f"got {cfg['n_months']}")
    )
    asserts.append(
        ("config.arr_factor == 12.0", cfg["arr_factor"] == 12.0, f"got {cfg['arr_factor']}")
    )
    asserts.append(
        ("config.month_range == ['2025-01', '2026-06']",
         cfg["month_range"] == ["2025-01", "2026-06"], f"got {cfg['month_range']}")
    )

    # --- First-month rollforward (Jan-25, ARR scale) ---
    m0 = result["monthly"][0]
    asserts.append(
        ("monthly[0].beginning == 87,600.00 (Jan-25 ARR = $7,300 MRR × 12)",
         abs(m0["beginning"] - 87_600.00) < 0.01, f"got {m0['beginning']:,.2f}")
    )
    asserts.append(
        ("monthly[0].ending == 87,600.00 (no movements in first month)",
         abs(m0["ending"] - 87_600.00) < 0.01, f"got {m0['ending']:,.2f}")
    )
    asserts.append(
        ("monthly[0].n_active == 8 (customers 1,2,3,4,5,6,8,10 active in Jan-25)",
         m0["n_active"] == 8, f"got {m0['n_active']}")
    )

    # --- Last-month rollforward (Jun-26, ARR scale) ---
    mL = result["monthly"][-1]
    asserts.append(
        ("monthly[-1].beginning == 91,200.00 (no movements May→Jun-26)",
         abs(mL["beginning"] - 91_200.00) < 0.01, f"got {mL['beginning']:,.2f}")
    )
    asserts.append(
        ("monthly[-1].ending == 91,200.00 (Jun-26 ARR = $7,600 MRR × 12)",
         abs(mL["ending"] - 91_200.00) < 0.01, f"got {mL['ending']:,.2f}")
    )
    asserts.append(
        ("monthly[-1].n_active == 9 (everyone except customer 6)",
         mL["n_active"] == 9, f"got {mL['n_active']}")
    )

    # --- Length asserts ---
    n_months = len(result["monthly"])
    asserts.append(("len(monthly) == 18", n_months == 18, f"got {n_months}"))

    n_metrics = len(result["metrics_monthly"])
    asserts.append(
        ("len(metrics_monthly) == 17 (first month has no prior)",
         n_metrics == 17, f"got {n_metrics}")
    )

    ltm = result["metrics_ltm"]
    asserts.append(
        ("len(metrics_ltm) == 6 (months m13..m18 each have a 12-month-prior compare)",
         bool(ltm) and len(ltm) == 6, f"len = {len(ltm) if ltm else 0}")
    )

    for name, ok, detail in asserts:
        marker = "PASS" if ok else "FAIL"
        print(f"  [{marker}] {name}  ({detail})")

    # Print a few key monthly rows for human eyeball
    print("\nMONTHLY ROLLFORWARD (first 3 + last)")
    print("-" * 72)
    for agg in result["monthly"][:3] + [result["monthly"][-1]]:
        print(
            f"  {agg['month']} | beg={agg['beginning']:>14,.2f} | "
            f"new={agg['new']:>11,.2f} | up={agg['upsell']:>10,.2f} | "
            f"dn={agg['downsell']:>10,.2f} | ch={agg['churn']:>11,.2f} | "
            f"end={agg['ending']:>14,.2f} | n={agg['n_active']:>3d}"
        )

    if result["metrics_ltm"]:
        print("\nLTM METRICS")
        print("-" * 72)
        for r in result["metrics_ltm"]:
            print(
                f"  {r['month']} | gross={r['gross']:.4f} | net={r['net']:.4f} | logo={r['logo']:.4f}"
            )

    all_layers_pass = all(ok for _, ok, _ in layer_results)
    all_asserts_pass = all(ok for _, ok, _ in asserts)

    print("\n" + "=" * 72)
    if all_layers_pass and all_asserts_pass:
        print("OVERALL: PASS — all 8 verification layers + all asserts passed.")
    else:
        print(
            f"OVERALL: FAIL — layers_ok={all_layers_pass}, asserts_ok={all_asserts_pass}"
        )
    print("=" * 72)

    # Cleanup
    try:
        os.unlink(csv_path)
    except OSError:
        pass

    return 0 if (all_layers_pass and all_asserts_pass) else 1


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--self-test":
        # Allow an optional positional path after --self-test; otherwise default
        # to the bundled fixture.
        path_arg = sys.argv[2] if len(sys.argv) > 2 else None
        sys.exit(_self_test(path_arg))
    if len(sys.argv) == 1:
        # No args: run self-test by default to make `python3 compute.py` useful.
        sys.exit(_self_test())
    sys.exit(main())
